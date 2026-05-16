"""
Pure-Python CPAN daily report generator.
No Excel master template. No formula dependency.
Reads 5 input files → builds formatted xlsx → exports PDF.
Also generates DL-DOWN-RT sheet from Down-DL-list + DL-alarms CSVs.
"""

import os
import re
from datetime import datetime
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ─────────────────────────────────────────────────────────────────────
_TITLE_FILL  = PatternFill("solid", fgColor="1F3864")
_HDR_FILL    = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TITLE_FONT  = Font(bold=True, color="FFFFFF", size=11)
_HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
_DATA_FONT   = Font(size=9)
_THIN        = Side(style="thin",   color="B0B0B0")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Node filter ─────────────────────────────────────────────────────────────────
# Primary filter: IP-based against NODE-LIST (GJ + WTR-GJ managed nodes).
# Fallback: CIRCLE column filter when NODE-LIST is not available.
_CIRCLES = {'GJ', 'WTR'}   # fallback circle filter (used only if no node_list.csv)

def _load_managed_ips():
    """Load managed node IPs from node_list.csv if present in project dir."""
    csv_path = os.path.join(os.path.dirname(__file__), 'node_list.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, dtype=str)
        col = df.columns[0]
        ips = set(df[col].dropna().str.strip().tolist())
        return ips if ips else None
    except Exception:
        return None

_MANAGED_IPS = _load_managed_ips()   # loaded once at import time

# ── SSA → BA (Billing Area) mapping from SSACODE master table ──────────────────
_SSA_BA = {
    # Gujarat (GJ) — maps SSA code → BA name
    # BA: Ahmedabad
    'SSA AM':  'Ahmedabad',
    'SSA NAD': 'Ahmedabad',
    # BA: Amreli
    'SSA AMR': 'Amreli',
    'SSA JND': 'Amreli',
    # BA: Bhavnagar
    'SSA BV':  'Bhavnagar',
    'SSA SEN': 'Bhavnagar',
    # BA: Bhuj
    'SSA BUJ': 'Bhuj',
    # BA: Mehsana
    'SSA MEH': 'Mehsana',
    'SSA HMR': 'Mehsana',
    'SSA PNP': 'Mehsana',
    # BA: Rajkot
    'SSA RJ':  'Rajkot',
    'SSA JMN': 'Rajkot',
    # BA: Surat
    'SSA SR':  'Surat',
    'SSA BCH': 'Surat',
    'SSA VAL': 'Surat',
    # BA: Vadodara
    'SSA VDR': 'Vadodara',
    'SSA GDH': 'Vadodara',
    # WTR — SSA codes used in WTR input files
    'AM':  'Ahmedabad',
    'RJ':  'Rajkot',
    'MBI': 'Mumbai',
    'NP':  'Nagpur',
    'JB':  'Jabalpur',
    'BPL': 'Bhopal',
    'RYP': 'Raipur',
}

def _lookup_ba(ssa_val):
    if not ssa_val or str(ssa_val).strip() in ('', 'nan', 'None'):
        return ''
    s = str(ssa_val).strip()
    return _SSA_BA.get(s, s)   # fall back to raw SSA value if not in map


# ── Node type normalisation (model → class code used in reference reports) ──────
_NODE_TYPE_MAP = {
    'TN725B': 'B2',
    'TN705B': 'B1',
    'TN703B': 'A1',
}

def _norm_type(val):
    s = str(val).strip() if val and str(val).strip() not in ('', 'nan', 'None') else ''
    return _NODE_TYPE_MAP.get(s, s)


# ── Date helpers ───────────────────────────────────────────────────────────────
_MIN_VALID_YEAR = 2010  # rows with Create Time before this year are excluded

def _parse_ct(val):
    """Parse '\t2026/04/02,00:09:58' or other formats → datetime (or None)."""
    if not val or str(val).strip() in ('nan', 'None', ''):
        return None
    s = str(val).strip()
    for fmt in ('%Y/%m/%d,%H:%M:%S', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(s, fmt)
            # Reject clearly invalid/epoch dates
            if dt.year < _MIN_VALID_YEAR:
                return None
            return dt
        except ValueError:
            continue
    return None


def _total_days(create_dt, report_date):
    if create_dt is None:
        return None
    rd = (report_date if isinstance(report_date, datetime)
          else datetime(report_date.year, report_date.month, report_date.day))
    # Compare date parts only (strip time) so alarms on same calendar date = 1 day
    rd_date = rd.replace(hour=0, minute=0, second=0, microsecond=0)
    ct_date = create_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    return max(0, (rd_date - ct_date).days)


def _bucket(days):
    if days is None:  return ''
    if days <= 1:     return 'L1'
    if days <= 3:     return '1to3'
    if days <= 7:     return '3to7'
    if days <= 15:    return '7to15'
    return 'G15'


def _fix_days(df):
    """Convert Total Days float column to int where not null."""
    if 'Total Days' in df.columns:
        df['Total Days'] = df['Total Days'].apply(
            lambda x: int(x) if x is not None and str(x) not in ('', 'nan') else '')
    return df


# WTR SSAs that are in-scope for GJ circle (WTR-Ahmedabad and WTR-Rajkot rings)
_WTR_GJ_SSAS = {'AM', 'RJ'}

def _filter_nodes(df, ip_col='Node IP', circle_col='Circle', ssa_col=None):
    """
    Filter to GJ + WTR-GJ managed nodes only.
    - All GJ rows are kept.
    - WTR rows only kept if SSA/REGION is AM or RJ (WTR-Ahmedabad / WTR-Rajkot rings).
    - If neither circle col nor SSA col present, return df unchanged.
    """
    if circle_col not in df.columns:
        return df

    gj_mask  = df[circle_col].astype(str).str.upper() == 'GJ'
    wtr_mask = df[circle_col].astype(str).str.upper() == 'WTR'

    if ssa_col and ssa_col in df.columns:
        wtr_gj_mask = wtr_mask & df[ssa_col].astype(str).str.strip().isin(_WTR_GJ_SSAS)
    else:
        # Fallback: include all WTR (if we can't check SSA)
        wtr_gj_mask = wtr_mask

    return df[gj_mask | wtr_gj_mask].copy()


def _load_dl_ems_ips():
    """Load DL-EMS managed link IPs from dl_ems_ips.csv if present."""
    csv_path = os.path.join(os.path.dirname(__file__), 'dl_ems_ips.csv')
    if not os.path.exists(csv_path):
        return None
    try:
        df = pd.read_csv(csv_path, dtype=str)
        ips = set(df.iloc[:, 0].dropna().str.strip().tolist())
        return ips if ips else None
    except Exception:
        return None

_DL_EMS_IPS = _load_dl_ems_ips()


def _filter_dl_nodes(df, aip_col='IP A End', zip_col='IP Z End'):
    """Filter DL-FAIL rows where A-end OR Z-end IP is a GJ/WTR-GJ managed link node."""
    if _DL_EMS_IPS and aip_col in df.columns and zip_col in df.columns:
        a_in = df[aip_col].astype(str).str.strip().isin(_DL_EMS_IPS)
        z_in = df[zip_col].astype(str).str.strip().isin(_DL_EMS_IPS)
        return df[a_in | z_in].copy()
    # Fallback: GJ circle filter only
    if 'Circle' in df.columns:
        return df[df['Circle'].astype(str).str.upper() == 'GJ'].copy()
    return df


# ── Input readers ──────────────────────────────────────────────────────────────
def read_card_off(path, report_date):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    cols = list(df.columns)

    # Format A (older): Description, IP, NAME, TYPE, CIRCLE, SSA, CARD, Create Time
    # Format B (newer, 07.05.26+): SR.NO., IP, NAME, TYPE, CIRCLE, REGION, PHASE, Create Time
    if 'CARD' in cols:
        df = df.rename(columns={
            'IP': 'Node IP', 'NAME': 'Node Name', 'TYPE': 'Type',
            'CIRCLE': 'Circle', 'SSA': 'SSA', 'CARD': 'Card',
            'Create Time': 'Create Time',
        })
        ssa_col = 'SSA'
        out_cols = ['Sr', 'Node IP', 'Node Name', 'Type', 'Circle', 'SSA',
                    'Card', 'Create Time', 'Total Days', 'BA', 'Bucket']
    else:
        # Format B (07.05.26+): REGION→SSA, PHASE→Card (standardise to format A names)
        df = df.rename(columns={
            'IP': 'Node IP', 'NAME': 'Node Name', 'TYPE': 'Type',
            'CIRCLE': 'Circle', 'REGION': 'SSA', 'PHASE': 'Card',
            'Create Time': 'Create Time',
        })
        ssa_col = 'SSA'
        out_cols = ['Sr', 'Node IP', 'Node Name', 'Type', 'Circle', 'SSA',
                    'Card', 'Create Time', 'Total Days', 'BA', 'Bucket']

    df['Create Time'] = df['Create Time'].str.strip()
    if 'Type' in df.columns:
        df['Type'] = df['Type'].apply(_norm_type)
    df['_dt']         = df['Create Time'].apply(_parse_ct)
    df = df.dropna(subset=['_dt'])                          # drops unparseable AND year<2010
    df = _filter_nodes(df, ip_col='Node IP', circle_col='Circle', ssa_col=ssa_col)
    df['Total Days']  = df['_dt'].apply(lambda x: _total_days(x, report_date))
    df['Bucket']      = df['Total Days'].apply(_bucket)
    df['BA']          = df[ssa_col].apply(_lookup_ba) if ssa_col in df.columns else ''
    df = df.sort_values('Total Days', ascending=False)
    df = _fix_days(df)
    df.insert(0, 'Sr', range(1, len(df) + 1))
    return df[[c for c in out_cols if c in df.columns]]


def read_device_off(path, report_date, dash_df=None):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={
        'IP': 'Node IP', 'NAME': 'Node Name', 'TYPE': 'Type',
        'CIRCLE': 'Circle', 'REGION': 'Region', 'PHASE': 'Phase',
        'Create Time': 'Create Time',
    })
    df['Create Time'] = df['Create Time'].str.strip()
    if 'Type' in df.columns:
        df['Type'] = df['Type'].apply(_norm_type)
    df['_dt']         = df['Create Time'].apply(_parse_ct)
    df = df.dropna(subset=['_dt'])
    df = _filter_nodes(df, ip_col='Node IP', circle_col='Circle', ssa_col='Region')
    df['Total Days']  = df['_dt'].apply(lambda x: _total_days(x, report_date))
    df['Bucket']      = df['Total Days'].apply(_bucket)
    df['BA']          = df['Region'].apply(_lookup_ba) if 'Region' in df.columns else ''
    df = df.sort_values('Total Days', ascending=False)
    df.insert(0, 'Sr', range(1, len(df) + 1))

    # Lookup Reason from DASH-DOWN by matching IP → Remarks
    df['Reason'] = ''
    if dash_df is not None:
        ip_col  = next((c for c in dash_df.columns if c.upper() == 'IP'), None)
        rem_col = next((c for c in dash_df.columns
                        if any(k in c.upper() for k in ('REMARK', 'REASON'))), None)
        if ip_col and rem_col:
            ip_map = dict(zip(
                dash_df[ip_col].astype(str).str.strip(),
                dash_df[rem_col].astype(str).str.strip()
            ))
            df['Reason'] = (df['Node IP'].astype(str).map(ip_map)
                            .fillna('').replace({'nan': '', 'NaN': '', 'None': ''}))

    # Clean up Reason: remove any nan/None/empty artefacts
    if 'Reason' in df.columns:
        df['Reason'] = df['Reason'].apply(
            lambda x: '' if (x is None or str(x).strip().lower() in ('nan', 'none', ''))
                      else str(x).strip()
        )
    want = ['Sr', 'Node IP', 'Node Name', 'Type', 'Circle', 'Region',
            'Create Time', 'Total Days', 'Reason', 'BA', 'Bucket']
    df = _fix_days(df)
    return df[[c for c in want if c in df.columns]]


def read_fan_fail(path, report_date):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={
        'IP': 'Node IP', 'Name': 'Node Name', 'Type': 'Type',
        'Circle': 'Circle', 'SSA': 'SSA', 'PHASE': 'Phase',
        'Alarm ID': 'Alarm ID', 'Create Time': 'Create Time',
    })
    df['Create Time'] = df['Create Time'].str.strip()
    if 'Alarm ID' in df.columns:
        df['Alarm ID'] = df['Alarm ID'].str.strip()   # remove leading \t tab
    if 'Type' in df.columns:
        df['Type'] = df['Type'].apply(_norm_type)
    df['_dt']        = df['Create Time'].apply(_parse_ct)
    df = df.dropna(subset=['_dt'])
    df = _filter_nodes(df, ip_col='Node IP', circle_col='Circle', ssa_col='SSA')
    df['Total Days'] = df['_dt'].apply(lambda x: _total_days(x, report_date))
    df['Bucket']     = df['Total Days'].apply(_bucket)
    df['BA']         = df['SSA'].apply(_lookup_ba) if 'SSA' in df.columns else ''
    df = df.sort_values('Total Days', ascending=False)
    df.insert(0, 'Sr', range(1, len(df) + 1))
    want = ['Sr', 'Node IP', 'Node Name', 'Type', 'Circle', 'SSA',
            'Alarm ID', 'Create Time', 'Total Days', 'BA', 'Bucket']
    df = _fix_days(df)
    return df[[c for c in want if c in df.columns]]


def read_dl_fail(path, report_date, links_df=None):
    df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={
        ' Circle': 'Circle', 'Circle': 'Circle', 'region': 'Region',
        'IP A END': 'IP A End', 'A END': 'A End',
        'IP Z END': 'IP Z End', 'Z END ': 'Z End', 'Z END': 'Z End',
        'Alarm Status': 'Alarm', 'Create Time': 'Create Time',
    })
    # Strip node-type prefix (TN725B_, TN705B_, etc.) from link name columns
    _tn_pfx = re.compile(r'^TN\d+[A-Z]+[\s_]+', re.IGNORECASE)
    for _col in ('A End', 'Z End'):
        if _col in df.columns:
            df[_col] = df[_col].apply(
                lambda v: _tn_pfx.sub('', str(v).strip()) if isinstance(v, str) else v
            )
    df['Create Time'] = df['Create Time'].str.strip()
    df['_dt']         = df['Create Time'].apply(_parse_ct)
    df = df.dropna(subset=['_dt'])
    df = _filter_dl_nodes(df, aip_col='IP A End', zip_col='IP Z End')
    df['Total Days']  = df['_dt'].apply(lambda x: _total_days(x, report_date))
    df['Bucket']      = df['Total Days'].apply(_bucket)
    df = df.sort_values('Total Days', ascending=False)
    df.insert(0, 'Sr', range(1, len(df) + 1))

    # Optional: lookup Bandwidth and CIR from network links data
    df['Bandwidth'] = ''
    df['CIR %']     = ''
    if links_df is not None:
        try:
            ldf = links_df.copy()
            ldf['_aip'] = ldf['A End'].astype(str).str.split('_').str[0].str.strip()
            ldf['_zip'] = ldf['Z End'].astype(str).str.split('_').str[0].str.strip()
            bw_map, cir_map = {}, {}
            for _, row in ldf.iterrows():
                key = tuple(sorted([row['_aip'], row['_zip']]))
                bw_map[key]  = row.get('Bandwidth', '')
                cir_map[key] = row.get('CIR Utilization Ratio(%)', '')

            def _lookup(r, mapping):
                k = tuple(sorted([
                    str(r.get('IP A End', '')).strip(),
                    str(r.get('IP Z End', '')).strip()
                ]))
                return mapping.get(k, '')

            df['Bandwidth'] = df.apply(lambda r: _lookup(r, bw_map),  axis=1)
            df['CIR %']     = df.apply(lambda r: _lookup(r, cir_map), axis=1)
        except Exception:
            pass

    want = ['Sr', 'Circle', 'Region', 'IP A End', 'A End', 'IP Z End', 'Z End',
            'Create Time', 'Total Days', 'Bandwidth', 'CIR %', 'Alarm', 'Bucket']
    df = _fix_days(df)
    return df[[c for c in want if c in df.columns]]


def read_dash_down(path):
    """Read DASH-DOWN CSV (10-col header, 13+ actual data cols) with forward-fill SSA grouping."""
    if path.lower().endswith('.csv'):
        with open(path, encoding='latin1') as fh:
            first_line = fh.readline()
        n_header = len(first_line.split(','))
        base_cols = ['SSA', 'SDCA', 'Exchange Code', 'Location', 'NE Type',
                     'IP', 'Down Time', 'Remarks', 'Down Days', 'BB Connection']
        extra = [f'_x{i}' for i in range(max(0, 20 - len(base_cols)))]
        all_names = (base_cols + extra)[:20]
        df = pd.read_csv(path, dtype=str, encoding='latin1',
                         names=all_names, skiprows=1,
                         on_bad_lines='warn')
    else:
        df = pd.read_excel(path, dtype=str)
    df.columns = [c.strip() for c in df.columns]

    # Forward-fill SSA/SDCA (group headers appear only on first row of each group)
    for col in ('SSA', 'SDCA'):
        if col in df.columns:
            df[col] = df[col].replace(
                {'nan': None, 'N/A': None, '#N/A': None, 'n/a': None, '': None}
            ).ffill()

    # Keep only rows with a valid IP address
    if 'IP' in df.columns:
        df = df[df['IP'].notna() &
                df['IP'].str.strip().str.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')]

    df = df.reset_index(drop=True)
    return df


# ── SUM-TABLE ─────────────────────────────────────────────────────────────────
_BUCKETS = ['L1', '1to3', '3to7', '7to15', 'G15']

# Fixed Gujarat BA → [(ssa_code, display_name), ...] mapping  (matches master template)
_GJ_STRUCTURE = [
    ('Ahmedabad', [('SSA AM',  'Ahmedabad'),  ('SSA NAD', 'Nadiad')]),
    ('Amreli',    [('SSA AMR', 'Amreli'),     ('SSA JND', 'Junagadh')]),
    ('Bhavnagar', [('SSA BV',  'Bhavnagar'),  ('SSA SEN', 'Surendranagar')]),
    ('Bhuj',      [('SSA BUJ', 'Bhuj')]),
    ('Mehsana',   [('SSA MEH', 'Mehsana'),    ('SSA HMR', 'Himatnagar'), ('SSA PNP', 'Palanpur')]),
    ('Rajkot',    [('SSA RJ',  'Rajkot'),     ('SSA JMN', 'Jamnagar')]),
    ('Surat',     [('SSA SR',  'Surat'),      ('SSA BCH', 'Bharuch'),    ('SSA VAL', 'Valsad')]),
    ('Vadodara',  [('SSA VDR', 'Vadodara'),   ('SSA GDH', 'Godhra')]),
]
_WTR_SSAS = [('AM', 'Ahmedabad'), ('RJ', 'Rajkot')]

# SSA col name per sheet type
_SHEET_SSA_COL = {
    'CARD-OFF-R':   'SSA',
    'DEVICE-OFF-R': 'Region',
    'FAN-FAIL-R':   'SSA',
    'DL-FAIL-R':    'Region',
}

_SECTION_TITLES = {
    'CARD-OFF-R':   'CARD OFFLINE REPORT  SUMMARY',
    'DEVICE-OFF-R': 'DEVICE OFFLINE REPORT  SUMMARY',
    'FAN-FAIL-R':   'FAN FAILURE REPORT  SUMMARY',
    'DL-FAIL-R':    'DL FAIL REPORT  SUMMARY',
}
_SHEETS_ORDER = ['CARD-OFF-R', 'DEVICE-OFF-R', 'FAN-FAIL-R', 'DL-FAIL-R']


# ── xlsx sheet writer ──────────────────────────────────────────────────────────
def _write_sheet(wb, tab_name, title, headers, data_rows, date_str):
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_properties.tabColor = "00B050"   # green tab
    n_cols = len(headers)

    # Row 1: title banner
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws['A1']
    c.value = f"DATE : {date_str}    {title}"
    c.font, c.fill, c.alignment = _TITLE_FONT, _TITLE_FILL, _CENTER
    ws.row_dimensions[1].height = 24

    # Row 2: column headers
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font, c.fill, c.alignment, c.border = _HDR_FONT, _HDR_FILL, _CENTER, _BORDER
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, row in enumerate(data_rows, 3):
        fill = _ALT_FILL if ri % 2 == 0 else _WHITE_FILL
        for ci in range(1, n_cols + 1):
            val = row[ci - 1] if ci - 1 < len(row) else ''
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = _DATA_FONT
            c.fill   = fill
            c.alignment = _CENTER if ci <= 2 else _LEFT
            c.border = _BORDER

    # Auto column width
    for ci, hdr in enumerate(headers, 1):
        col_vals = [str(hdr)] + [str(r[ci - 1]) for r in data_rows if ci - 1 < len(r)]
        width = min(max((len(v) for v in col_vals if v), default=8), 45)
        ws.column_dimensions[get_column_letter(ci)].width = width + 2

    ws.freeze_panes = "A3"

    # Print/PDF page setup — landscape A4, fit all columns on one page width
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = 9        # 9 = A4
    ws.page_setup.fitToWidth  = 1        # all columns on 1 page wide
    ws.page_setup.fitToHeight = 0        # unlimited pages tall
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    return ws


def _write_sum_table(wb, sheet_dfs, date_str):
    """Write SUM-TABLE matching CPAN master template — fixed BA→SSA, Gujarat+WTR+GrandTotal."""
    ws = wb.create_sheet(title='SUM-TABLE', index=0)
    ws.sheet_properties.tabColor = "FFC000"

    # ── Styles ────────────────────────────────────────────────────────────────
    CYAN_FILL   = PatternFill("solid", fgColor="00FFFF")
    BLUE_FILL   = PatternFill("solid", fgColor="1F3864")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
    TOTAL_FILL  = PatternFill("solid", fgColor="FCE4D6")   # light salmon — Gujarat/WTR total
    GRAND_FILL  = PatternFill("solid", fgColor="FFC000")   # amber — Grand Total
    WTR_FILL    = PatternFill("solid", fgColor="DDEBF7")   # light blue — WTR rows

    TITLE_FONT  = Font(name='Calibri', bold=True, size=12)
    HDR_FONT    = Font(name='Calibri', bold=True, size=10, color="FFFFFF")
    DATA_FONT   = Font(name='Calibri', size=10)
    BOLD_FONT   = Font(name='Calibri', bold=True, size=10)
    TOTAL_FONT  = Font(name='Calibri', bold=True, size=10)

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    THIN   = Side(style='thin')
    BDR    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    N = 9  # columns

    # Column widths
    for ci, w in enumerate([12, 15, 18, 12, 12, 12, 12, 12, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    col_headers = [
        'Circle', 'BA', 'SSA',
        'Fault \nPending \nfor\n< 1 DAY',
        'Fault \nPending \nfor\n> 1 - 3 DAY',
        'Fault \nPending \nfor\n> 3 - 7 DAY',
        'Fault \nPending \nfor\n> 7 - 15 DAY',
        'Fault \nPending \nfor\n> 15 DAY',
        'Total\nFault \nPending ',
    ]

    def _counts(df, ssa_col, code):
        """Return list of 5 bucket counts for a given SSA code."""
        if df is None or ssa_col not in df.columns or 'Bucket' not in df.columns:
            return [0] * 5
        sub = df[df[ssa_col].astype(str).str.strip() == code]
        return [int((sub['Bucket'] == b).sum()) for b in _BUCKETS]

    def _put(r, c, val, font=DATA_FONT, fill=WHITE_FILL, align=CENTER, border=BDR):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font, cell.fill, cell.alignment, cell.border = font, fill, align, border
        return cell

    def _merge(r1, c1, r2, c2, val, font=DATA_FONT, fill=WHITE_FILL):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=val)
        cell.font, cell.fill, cell.alignment = font, fill, CENTER
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
            for cc in row:
                cc.border = BDR
        return cell

    def _total_row(r, label, counts, fill):
        _merge(r, 1, r, 3, label, TOTAL_FONT, fill)
        for ci, cnt in enumerate(counts, 4):
            _put(r, ci, cnt, TOTAL_FONT, fill)
        _put(r, 9, sum(counts), TOTAL_FONT, fill)
        ws.row_dimensions[r].height = 18

    row = 1

    for sh_name in _SHEETS_ORDER:
        df      = sheet_dfs.get(sh_name)
        ssa_col = _SHEET_SSA_COL.get(sh_name, 'SSA')
        title   = _SECTION_TITLES.get(sh_name, sh_name)

        # ── Title banner ──────────────────────────────────────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
        c = ws.cell(row=row, column=1,
                    value=f"  DATE : {date_str}                {title}           Gujarat Circle")
        c.font, c.fill, c.alignment = TITLE_FONT, CYAN_FILL, CENTER
        ws.row_dimensions[row].height = 28
        row += 1
        row += 1  # blank spacer

        # ── Column headers ────────────────────────────────────────────────────
        for ci, hdr in enumerate(col_headers, 1):
            _put(row, ci, hdr, HDR_FONT, BLUE_FILL)
        ws.row_dimensions[row].height = 60
        row += 1

        # ── Gujarat data ──────────────────────────────────────────────────────
        gj_start = row
        gj_total = [0] * 5

        for ba_name, ssa_list in _GJ_STRUCTURE:
            ba_start = row
            for ssa_idx, (ssa_code, ssa_display) in enumerate(ssa_list):
                cnts  = _counts(df, ssa_col, ssa_code)
                gj_total = [gj_total[i] + cnts[i] for i in range(5)]
                fill  = WHITE_FILL if (row % 2 == 0) else ALT_FILL
                _put(row, 3, ssa_display, DATA_FONT, fill)
                for ci, cnt in enumerate(cnts, 4):
                    _put(row, ci, cnt, DATA_FONT, fill)
                _put(row, 9, sum(cnts), BOLD_FONT, fill)
                ws.row_dimensions[row].height = 18
                row += 1

            ba_end = row - 1
            if ba_end > ba_start:
                _merge(ba_start, 2, ba_end, 2, ba_name, BOLD_FONT, WHITE_FILL)
            else:
                _put(ba_start, 2, ba_name, BOLD_FONT, WHITE_FILL)

        gj_end = row - 1
        if gj_end > gj_start:
            _merge(gj_start, 1, gj_end, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
        else:
            _put(gj_start, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)

        # Gujarat Total
        _total_row(row, 'Gujarat Total', gj_total, TOTAL_FILL)
        row += 1
        row += 1  # blank

        # ── WTR data ──────────────────────────────────────────────────────────
        wtr_start  = row
        wtr_total  = [0] * 5

        for wtr_idx, (ssa_code, ssa_display) in enumerate(_WTR_SSAS):
            cnts = _counts(df, ssa_col, ssa_code)
            wtr_total = [wtr_total[i] + cnts[i] for i in range(5)]
            fill = WTR_FILL if wtr_idx % 2 == 0 else ALT_FILL
            _put(row, 3, ssa_display, DATA_FONT, fill)
            for ci, cnt in enumerate(cnts, 4):
                _put(row, ci, cnt, DATA_FONT, fill)
            _put(row, 9, sum(cnts), BOLD_FONT, fill)
            ws.row_dimensions[row].height = 18
            row += 1

        wtr_end = row - 1
        if wtr_end > wtr_start:
            _merge(wtr_start, 1, wtr_end, 1, 'WTR', BOLD_FONT, WTR_FILL)
            _merge(wtr_start, 2, wtr_end, 2, 'WTR', BOLD_FONT, WTR_FILL)
        else:
            _put(wtr_start, 1, 'WTR', BOLD_FONT, WTR_FILL)
            _put(wtr_start, 2, 'WTR', BOLD_FONT, WTR_FILL)

        # WTR Total
        _total_row(row, 'WTR  Total', wtr_total, TOTAL_FILL)
        row += 1
        row += 1  # blank

        # Grand Total
        grand = [gj_total[i] + wtr_total[i] for i in range(5)]
        _total_row(row, 'Grand Total', grand, GRAND_FILL)
        row += 1

        row += 4  # gap before next section

    # ── Section 5: GCS PORTAL BOOKED CPAN DOCKET SUMMARY ─────────────────────
    DOCKET_COLS = [
        'Circle', 'BA', 'SSA',
        'No. of\nDockets\nBooked', 'Resolved', 'Closed', 'assigned', 'feedback', '',
    ]
    ND = len(DOCKET_COLS)
    PEACH_HDR_FILL = PatternFill("solid", fgColor="FCE4D6")

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ND)
    c = ws.cell(row=row, column=1, value="    GCS PORTAL  - BOOKED CPAN DOCKET SUMMARY ")
    c.font, c.fill, c.alignment = TITLE_FONT, CYAN_FILL, CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    row += 1  # blank

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ND)
    c = ws.cell(row=row, column=1, value="From Date :                         TO  Date :")
    c.font, c.fill, c.alignment = TITLE_FONT, CYAN_FILL, CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    # Docket column headers (peach background)
    for ci, hdr in enumerate(DOCKET_COLS, 1):
        _put(row, ci, hdr, BOLD_FONT, PEACH_HDR_FILL)
    ws.row_dimensions[row].height = 50
    row += 1

    # Gujarat data — all zeros (no docket input available)
    gj_doc_start = row
    for ba_name, ssa_list in _GJ_STRUCTURE:
        ba_start = row
        for ssa_idx, (ssa_code, ssa_display) in enumerate(ssa_list):
            fill = WHITE_FILL if row % 2 == 0 else ALT_FILL
            _put(row, 3, ssa_display, DATA_FONT, fill)
            for ci in range(4, ND + 1):
                _put(row, ci, 0, DATA_FONT, fill)
            ws.row_dimensions[row].height = 18
            row += 1
        ba_end = row - 1
        if ba_end > ba_start:
            _merge(ba_start, 2, ba_end, 2, ba_name, BOLD_FONT, WHITE_FILL)
        else:
            _put(ba_start, 2, ba_name, BOLD_FONT, WHITE_FILL)
    gj_doc_end = row - 1
    if gj_doc_end > gj_doc_start:
        _merge(gj_doc_start, 1, gj_doc_end, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
    else:
        _put(gj_doc_start, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)

    # Gujarat Total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    c = ws.cell(row=row, column=1, value='Gujarat Total')
    c.font, c.fill, c.alignment = TOTAL_FONT, TOTAL_FILL, CENTER
    for r2 in range(1, 4):
        ws.cell(row, r2).border = BDR
        ws.cell(row, r2).fill  = TOTAL_FILL
    for ci in range(4, ND + 1):
        _put(row, ci, 0, TOTAL_FONT, TOTAL_FILL)
    ws.row_dimensions[row].height = 18
    row += 1

    row += 4  # gap

    # ── Section 6: DEGRADED DL REPORT SUMMARY ────────────────────────────────
    # Use DEGR-DL-R data if available, else all zeros
    degr_df = sheet_dfs.get('DEGR-DL-R')
    degr_ssa_col = 'Region'

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    c = ws.cell(row=row, column=1,
                value=f"  DATE : {date_str}                DEGRADED DL REPORT  SUMMARY           Gujarat Circle")
    c.font, c.fill, c.alignment = TITLE_FONT, CYAN_FILL, CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    row += 1  # blank

    for ci, hdr in enumerate(col_headers, 1):
        _put(row, ci, hdr, HDR_FONT, BLUE_FILL)
    ws.row_dimensions[row].height = 60
    row += 1

    # Gujarat data
    gj_dg_start = row
    gj_dg_total = [0] * 5
    for ba_name, ssa_list in _GJ_STRUCTURE:
        ba_start = row
        for ssa_idx, (ssa_code, ssa_display) in enumerate(ssa_list):
            cnts = _counts(degr_df, degr_ssa_col, ssa_code) if degr_df is not None else [0] * 5
            gj_dg_total = [gj_dg_total[i] + cnts[i] for i in range(5)]
            fill = WHITE_FILL if row % 2 == 0 else ALT_FILL
            _put(row, 3, ssa_display, DATA_FONT, fill)
            for ci, cnt in enumerate(cnts, 4):
                _put(row, ci, cnt, DATA_FONT, fill)
            _put(row, 9, sum(cnts), BOLD_FONT, fill)
            ws.row_dimensions[row].height = 18
            row += 1
        ba_end = row - 1
        if ba_end > ba_start:
            _merge(ba_start, 2, ba_end, 2, ba_name, BOLD_FONT, WHITE_FILL)
        else:
            _put(ba_start, 2, ba_name, BOLD_FONT, WHITE_FILL)
    gj_dg_end = row - 1
    if gj_dg_end > gj_dg_start:
        _merge(gj_dg_start, 1, gj_dg_end, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
    else:
        _put(gj_dg_start, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
    _total_row(row, 'Gujarat Total', gj_dg_total, TOTAL_FILL)
    row += 1
    row += 1

    # WTR
    wtr_dg_start = row
    wtr_dg_total = [0] * 5
    for wtr_idx, (ssa_code, ssa_display) in enumerate(_WTR_SSAS):
        cnts = _counts(degr_df, degr_ssa_col, ssa_code) if degr_df is not None else [0] * 5
        wtr_dg_total = [wtr_dg_total[i] + cnts[i] for i in range(5)]
        fill = WTR_FILL if wtr_idx % 2 == 0 else ALT_FILL
        _put(row, 3, ssa_display, DATA_FONT, fill)
        for ci, cnt in enumerate(cnts, 4):
            _put(row, ci, cnt, DATA_FONT, fill)
        _put(row, 9, sum(cnts), BOLD_FONT, fill)
        ws.row_dimensions[row].height = 18
        row += 1
    wtr_dg_end = row - 1
    if wtr_dg_end > wtr_dg_start:
        _merge(wtr_dg_start, 1, wtr_dg_end, 1, 'WTR', BOLD_FONT, WTR_FILL)
        _merge(wtr_dg_start, 2, wtr_dg_end, 2, 'WTR', BOLD_FONT, WTR_FILL)
    else:
        _put(wtr_dg_start, 1, 'WTR', BOLD_FONT, WTR_FILL)
        _put(wtr_dg_start, 2, 'WTR', BOLD_FONT, WTR_FILL)
    _total_row(row, 'WTR  Total', wtr_dg_total, TOTAL_FILL)
    row += 1
    row += 1
    grand_dg = [gj_dg_total[i] + wtr_dg_total[i] for i in range(5)]
    _total_row(row, 'Grand Total', grand_dg, GRAND_FILL)
    row += 1

    row += 4  # gap

    # ── Section 7: DEVICE OFFLINE REPORT SUMMARY - DASHBOARD ─────────────────
    LAVENDER_FILL = PatternFill("solid", fgColor="D9D2E9")
    dash_df  = sheet_dfs.get('DEVICE-OFF-R')   # same data as device offline
    dash_col = 'Region'

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N)
    c = ws.cell(row=row, column=1,
                value=f"  DATE : {date_str}                DEVICE OFFLINE REPORT  SUMMARY - DASHBOARD             Gujarat Circle")
    c.font, c.fill, c.alignment = TITLE_FONT, LAVENDER_FILL, CENTER
    ws.row_dimensions[row].height = 28
    row += 1
    row += 1  # blank

    for ci, hdr in enumerate(col_headers, 1):
        _put(row, ci, hdr, HDR_FONT, BLUE_FILL)
    ws.row_dimensions[row].height = 60
    row += 1

    gj_db_start = row
    gj_db_total = [0] * 5
    for ba_name, ssa_list in _GJ_STRUCTURE:
        ba_start = row
        for ssa_idx, (ssa_code, ssa_display) in enumerate(ssa_list):
            cnts = _counts(dash_df, dash_col, ssa_code)
            gj_db_total = [gj_db_total[i] + cnts[i] for i in range(5)]
            fill = WHITE_FILL if row % 2 == 0 else ALT_FILL
            _put(row, 3, ssa_display, DATA_FONT, fill)
            for ci, cnt in enumerate(cnts, 4):
                _put(row, ci, cnt, DATA_FONT, fill)
            _put(row, 9, sum(cnts), BOLD_FONT, fill)
            ws.row_dimensions[row].height = 18
            row += 1
        ba_end = row - 1
        if ba_end > ba_start:
            _merge(ba_start, 2, ba_end, 2, ba_name, BOLD_FONT, WHITE_FILL)
        else:
            _put(ba_start, 2, ba_name, BOLD_FONT, WHITE_FILL)
    gj_db_end = row - 1
    if gj_db_end > gj_db_start:
        _merge(gj_db_start, 1, gj_db_end, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
    else:
        _put(gj_db_start, 1, 'Gujarat', BOLD_FONT, WHITE_FILL)
    _total_row(row, 'Gujarat Total', gj_db_total, TOTAL_FILL)
    row += 1
    row += 1

    wtr_db_start = row
    wtr_db_total = [0] * 5
    for wtr_idx, (ssa_code, ssa_display) in enumerate(_WTR_SSAS):
        cnts = _counts(dash_df, dash_col, ssa_code)
        wtr_db_total = [wtr_db_total[i] + cnts[i] for i in range(5)]
        fill = WTR_FILL if wtr_idx % 2 == 0 else ALT_FILL
        _put(row, 3, ssa_display, DATA_FONT, fill)
        for ci, cnt in enumerate(cnts, 4):
            _put(row, ci, cnt, DATA_FONT, fill)
        _put(row, 9, sum(cnts), BOLD_FONT, fill)
        ws.row_dimensions[row].height = 18
        row += 1
    wtr_db_end = row - 1
    if wtr_db_end > wtr_db_start:
        _merge(wtr_db_start, 1, wtr_db_end, 1, 'WTR', BOLD_FONT, WTR_FILL)
        _merge(wtr_db_start, 2, wtr_db_end, 2, 'WTR', BOLD_FONT, WTR_FILL)
    else:
        _put(wtr_db_start, 1, 'WTR', BOLD_FONT, WTR_FILL)
        _put(wtr_db_start, 2, 'WTR', BOLD_FONT, WTR_FILL)
    _total_row(row, 'WTR  Total', wtr_db_total, TOTAL_FILL)
    row += 1
    row += 1
    grand_db = [gj_db_total[i] + wtr_db_total[i] for i in range(5)]
    _total_row(row, 'Grand Total', grand_db, GRAND_FILL)

    ws.freeze_panes = "A1"

    # Page setup for PDF
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


# ── DL Down Real-Time Report ───────────────────────────────────────────────────

def _clean(val):
    """Strip tabs, NBSP, and surrounding whitespace from a string value."""
    return str(val).replace('\t', '').replace('\xa0', ' ').strip()


def _extract_endpoint(raw):
    """
    Parse A-End / Z-End field like:
      10.121.18.245_SECTOR 11 GNE_TN725B_GJ_SSA AM_PH1\\S3P1
    Returns (ip, name, port).
    """
    s = _clean(raw)
    idx1 = s.find('_')
    if idx1 == -1:
        return s, '', ''
    ip = s[:idx1].strip()
    idx_tn7 = s.find('_TN7', idx1)
    if idx_tn7 != -1:
        name = s[idx1 + 1:idx_tn7].strip()
    else:
        idx2 = s.find('_', idx1 + 1)
        name = s[idx1 + 1:idx2].strip() if idx2 != -1 else s[idx1 + 1:].strip()
    idx_bs = s.find('\\')
    port = s[idx_bs + 1:].strip() if idx_bs != -1 else ''
    return ip, name, port


def read_dl_down_rt(down_dl_path, dl_alarms_path, report_date):
    """
    Replicates the logic of the CPAN REAL TIME DOWN DL Master.xlsx.

    Inputs
    ------
    down_dl_path  : path to Down-DL-list_dd-mm-yyyy.csv
    dl_alarms_path: path to DL-alarms_dd-mm-yyyy.csv
    report_date   : datetime

    Returns
    -------
    DataFrame with columns:
      Sr, SSA, DL Type, A End of DL, Z End of DL, Full DL Name,
      Down Date, Down Time, Down Days, A-END IP, A-END Name,
      Z-END IP, Z-END Name, OA-mix, DateTime-MIX, A-port, Z-port, Client
    """
    today = report_date if isinstance(report_date, datetime) else datetime(
        report_date.year, report_date.month, report_date.day)

    # ── Step 1: build alarm lookup from DL-alarms CSV ──────────────────────────
    # Key = cleaned DL name, Value = (down_date_str DD-MM-YYYY, time_str, full_dt_str)
    alarm_map = {}
    try:
        alarms_df = pd.read_csv(dl_alarms_path, dtype=str, encoding='utf-8-sig',
                                on_bad_lines='warn')
        alarms_df.columns = [c.strip() for c in alarms_df.columns]
        svc_col = next((c for c in alarms_df.columns
                        if 'SERVICE ID' in c.upper() or c.upper() == 'SERVICE ID'), None)
        ct_col  = next((c for c in alarms_df.columns
                        if 'CREATE TIME' in c.upper()), None)
        if svc_col and ct_col:
            for _, row in alarms_df.iterrows():
                svc = _clean(str(row.get(svc_col, '')))
                eq  = svc.find('=')
                if eq == -1:
                    continue
                dl_name = svc[eq + 1:].strip()
                if not dl_name or dl_name in alarm_map:
                    continue  # keep earliest occurrence
                ct_raw = _clean(str(row.get(ct_col, '')))
                # ct_raw e.g. "2026/05/15,16:23:40"
                comma = ct_raw.find(',')
                if comma != -1:
                    date_ymd = ct_raw[:comma].strip()   # 2026/05/15
                    time_str = ct_raw[comma + 1:].strip()
                    parts = date_ymd.split('/')
                    if len(parts) == 3:
                        date_dmy = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        date_dmy = date_ymd
                else:
                    date_ymd = ct_raw
                    time_str = ''
                    date_dmy = date_ymd
                alarm_map[dl_name] = (date_dmy, time_str, ct_raw)
    except Exception:
        pass

    # ── Step 2: process Down-DL-list CSV ──────────────────────────────────────
    try:
        dl_df = pd.read_csv(down_dl_path, dtype=str, encoding='utf-8-sig',
                            on_bad_lines='warn')
    except Exception:
        dl_df = pd.read_csv(down_dl_path, dtype=str, encoding='latin1',
                            on_bad_lines='warn')
    dl_df.columns = [c.strip() for c in dl_df.columns]

    rows = []
    for _, row in dl_df.iterrows():
        raw_name = _clean(str(row.get('Name', '')))
        if not raw_name:
            continue

        # SSA = part before first "-"
        dash = raw_name.find('-')
        ssa  = raw_name[:dash].strip() if dash != -1 else raw_name

        bw         = _clean(str(row.get('Bandwidth', '')))
        a_end_raw  = _clean(str(row.get('A End', '')))
        z_end_raw  = _clean(str(row.get('Z End', '')))
        client_val = _clean(str(row.get('Client', '')))

        a_ip, a_name, a_port = _extract_endpoint(a_end_raw)
        z_ip, z_name, z_port = _extract_endpoint(z_end_raw)

        a_mix = f"{a_ip}_{a_name}_{a_port}" if a_ip else ''
        z_mix = f"{z_ip}_{z_name}_{z_port}" if z_ip else ''

        # Alarm cross-reference
        alarm_info = alarm_map.get(raw_name, ('', '', ''))
        down_date, down_time, full_dt = alarm_info

        # Down Days = today − down_date
        down_days = ''
        if down_date:
            try:
                dd_parts = down_date.split('-')   # DD-MM-YYYY
                down_dt  = datetime(int(dd_parts[2]), int(dd_parts[1]), int(dd_parts[0]))
                diff     = (today.replace(hour=0, minute=0, second=0, microsecond=0) -
                            down_dt.replace(hour=0, minute=0, second=0, microsecond=0)).days
                down_days = max(0, diff)
            except Exception:
                pass

        rows.append({
            'SSA':         ssa,
            'DL Type':     bw,
            'A End of DL': a_mix,
            'Z End of DL': z_mix,
            'Full DL Name': raw_name,
            'Down Date':   down_date,
            'Down Time':   down_time,
            'Down Days':   down_days,
            'A-END IP':    a_ip,
            'A-END Name':  a_name,
            'Z-END IP':    z_ip,
            'Z-END Name':  z_name,
            'OA-mix':      ssa,
            'DateTime-MIX': full_dt,
            'A-port':      a_port,
            'Z-port':      z_port,
            'Client':      client_val,
        })

    if not rows:
        return pd.DataFrame(columns=[
            'Sr', 'SSA', 'DL Type', 'A End of DL', 'Z End of DL', 'Full DL Name',
            'Down Date', 'Down Time', 'Down Days', 'A-END IP', 'A-END Name',
            'Z-END IP', 'Z-END Name', 'OA-mix', 'DateTime-MIX', 'A-port', 'Z-port', 'Client'
        ])

    result = pd.DataFrame(rows)
    # Sort: DLs with known down date first (by down_days desc), then unknowns
    result['_sort'] = result['Down Days'].apply(
        lambda x: -int(x) if isinstance(x, int) else 1)
    result = result.sort_values('_sort').drop(columns=['_sort']).reset_index(drop=True)
    result.insert(0, 'Sr', range(1, len(result) + 1))
    return result


def _write_dl_down_rt_sheet(wb, df, date_str):
    """Write DL-DOWN-RT sheet to workbook with same styling as other report sheets."""
    _DL_DOWN_TITLE_FILL = PatternFill("solid", fgColor="1F3864")
    _DL_DOWN_HDR_FILL   = PatternFill("solid", fgColor="C00000")   # dark red tab
    _DL_DOWN_ALT_FILL   = PatternFill("solid", fgColor="FFE4E4")
    _DL_DOWN_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    ws = wb.create_sheet(title='DL-DOWN-RT')
    ws.sheet_properties.tabColor = "C00000"

    headers = list(df.columns)
    n_cols  = len(headers)

    # Row 1: title
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws['A1']
    c.value = (f"DATE : {date_str}    "
               f"Gujarat Circle  -  CPAN Real-Time Down DL Report  -  "
               f"Total : {len(df)} DLs")
    c.font      = _TITLE_FONT
    c.fill      = _DL_DOWN_TITLE_FILL
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 24

    # Row 2: headers
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = _HDR_FONT
        c.fill      = _DL_DOWN_HDR_FILL
        c.alignment = _CENTER
        c.border    = _BORDER
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, row_data in enumerate(df.values.tolist(), 3):
        fill = _DL_DOWN_ALT_FILL if ri % 2 == 0 else _DL_DOWN_WHITE_FILL
        for ci in range(1, n_cols + 1):
            val = row_data[ci - 1] if ci - 1 < len(row_data) else ''
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font      = _DATA_FONT
            c.fill      = fill
            c.alignment = _CENTER if ci <= 2 else _LEFT
            c.border    = _BORDER

    # Column widths
    for ci, hdr in enumerate(headers, 1):
        col_vals = [str(hdr)] + [str(r[ci - 1]) for r in df.values.tolist()
                                 if ci - 1 < len(r)]
        width = min(max((len(v) for v in col_vals if v), default=8), 50)
        ws.column_dimensions[get_column_letter(ci)].width = width + 2

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left   = 0.4
    ws.page_margins.right  = 0.4
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    return ws


# ── Main entry point ───────────────────────────────────────────────────────────
def generate_report(file_map, report_date, output_path, links_df=None):
    """
    file_map   : dict  {'card_off': path, 'device_off': path, 'fan_fail': path,
                         'dl_fail': path, 'dash_down': path}
    report_date: datetime
    output_path: str  — where to save the .xlsx
    links_df   : optional DataFrame from links.csv for DL bandwidth lookup
    Returns    : list of log messages
    """
    logs = []
    date_str = report_date.strftime('%d-%m-%Y')

    # Read DASH-DOWN first (needed for DEVICE-OFF reason lookup)
    dash_df = None
    if file_map.get('dash_down'):
        try:
            dash_df = read_dash_down(file_map['dash_down'])
            logs.append(f"  DASH-DOWN: {len(dash_df)} rows read")
        except Exception as e:
            logs.append(f"  DASH-DOWN: SKIP — {e}")

    wb = xl.Workbook()
    wb.remove(wb.active)

    sheet_dfs = {}   # collected for SUM-TABLE

    # CARD-OFF-R
    card_df = None
    if file_map.get('card_off'):
        try:
            card_df = read_card_off(file_map['card_off'], report_date)
            _write_sheet(wb, 'CARD-OFF-R', 'CARD OFFLINE REPORT',
                         list(card_df.columns), card_df.values.tolist(), date_str)
            sheet_dfs['CARD-OFF-R'] = card_df
            logs.append(f"  CARD-OFF-R: {len(card_df)} rows")
        except Exception as e:
            logs.append(f"  CARD-OFF-R: SKIP — {e}")

    # DEVICE-OFF-R
    dev_df = None
    if file_map.get('device_off'):
        try:
            dev_df = read_device_off(file_map['device_off'], report_date, dash_df)
            _write_sheet(wb, 'DEVICE-OFF-R', 'DEVICE OFFLINE REPORT',
                         list(dev_df.columns), dev_df.values.tolist(), date_str)
            sheet_dfs['DEVICE-OFF-R'] = dev_df
            logs.append(f"  DEVICE-OFF-R: {len(dev_df)} rows")
        except Exception as e:
            logs.append(f"  DEVICE-OFF-R: SKIP — {e}")

    # FAN-FAIL-R
    fan_df = None
    if file_map.get('fan_fail'):
        try:
            fan_df = read_fan_fail(file_map['fan_fail'], report_date)
            _write_sheet(wb, 'FAN-FAIL-R', 'FAN FAILURE REPORT',
                         list(fan_df.columns), fan_df.values.tolist(), date_str)
            sheet_dfs['FAN-FAIL-R'] = fan_df
            logs.append(f"  FAN-FAIL-R: {len(fan_df)} rows")
        except Exception as e:
            logs.append(f"  FAN-FAIL-R: SKIP — {e}")

    # DL-FAIL-R
    dl_df = None
    if file_map.get('dl_fail'):
        try:
            dl_df = read_dl_fail(file_map['dl_fail'], report_date, links_df)
            _write_sheet(wb, 'DL-FAIL-R', 'CPAN DL FAIL REPORT',
                         list(dl_df.columns), dl_df.values.tolist(), date_str)
            sheet_dfs['DL-FAIL-R'] = dl_df
            logs.append(f"  DL-FAIL-R: {len(dl_df)} rows")
        except Exception as e:
            logs.append(f"  DL-FAIL-R: SKIP — {e}")

    # DASH-DOWN-R (nodes down > 3 days)
    if dash_df is not None and len(dash_df):
        try:
            days_col = next((c for c in dash_df.columns
                             if 'DOWN' in c.upper() and 'DAY' in c.upper()), None)
            fdf = dash_df.copy()
            if days_col:
                fdf = fdf[pd.to_numeric(fdf[days_col], errors='coerce') > 3]
            fdf = fdf.reset_index(drop=True)
            fdf.insert(0, 'Sr', range(1, len(fdf) + 1))
            _write_sheet(wb, 'DASH-DOWN-R', 'CPAN NODE FAILURE REPORT (>3 Days)',
                         list(fdf.columns), fdf.values.tolist(), date_str)
            logs.append(f"  DASH-DOWN-R: {len(fdf)} rows")
        except Exception as e:
            logs.append(f"  DASH-DOWN-R: SKIP — {e}")

    # DL-DOWN-RT
    if file_map.get('dl_down') and file_map.get('dl_alarms'):
        try:
            dl_rt_df = read_dl_down_rt(
                file_map['dl_down'], file_map['dl_alarms'], report_date)
            _write_dl_down_rt_sheet(wb, dl_rt_df, date_str)
            logs.append(f"  DL-DOWN-RT: {len(dl_rt_df)} rows")
        except Exception as e:
            logs.append(f"  DL-DOWN-RT: SKIP — {e}")

    # SUM-TABLE (first sheet, index=0)
    if sheet_dfs:
        try:
            _write_sum_table(wb, sheet_dfs, date_str)
            logs.append(f"  SUM-TABLE: 4 sections")
        except Exception as e:
            logs.append(f"  SUM-TABLE: SKIP — {e}")

    wb.save(output_path)
    logs.append(f"xlsx saved: {os.path.basename(output_path)}")
    return logs
