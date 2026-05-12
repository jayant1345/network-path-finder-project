"""Deep comparison: reference vs input files for 02.04.26."""
import openpyxl
import pandas as pd

REF = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_02.04.26.xlsx"
INPUT_CARD = r"C:\Project_AI\cpan_report\CARD OFFLINE 02-04-2026.xlsx"
INPUT_DEV  = r"C:\Project_AI\cpan_report\DEVICE OFFLINE 02-04-2026.xlsx"
INPUT_FAN  = r"C:\Project_AI\cpan_report\FAN FAILURE 02-04-2026.xlsx"
INPUT_DL   = r"C:\Project_AI\cpan_report\CPAN DL FAIL REPORT  02-04-2026.xlsx"

def ref_df(wb, sh_name, header_row):
    sh = wb[sh_name]
    rows = [[c.value for c in sh[r]] for r in range(header_row, sh.max_row + 1)]
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df.columns = [str(c).strip().replace('\n',' ') if c else f"col{i}" for i, c in enumerate(df.columns)]
    return df

ref_wb = openpyxl.load_workbook(REF, data_only=True)

# ── CARD-OFF-R ─────────────────────────────────────────────────────────────
print("=" * 70)
print("CARD-OFF-R")
print("=" * 70)
card_in = pd.read_excel(INPUT_CARD)
ref_card = ref_df(ref_wb, "CARD-OFF-R", 4)

print(f"Input rows: {len(card_in)}, unique IPs: {card_in['IP'].nunique()}")
print(f"Ref rows:   {len(ref_card)}, unique IPs: {ref_card['Node IP'].dropna().nunique()}")
print(f"\nInput CIRCLE distribution:")
print(card_in['CIRCLE'].value_counts().to_string())
print(f"\nRef Circle distribution (col 5):")
print(ref_card['Circle'].value_counts().head(20).to_string())

print(f"\nInput SSA distribution (top 10):")
print(card_in['SSA'].value_counts().head(10).to_string())
print(f"\nRef SSA distribution (top 10):")
print(ref_card['SSA'].value_counts().head(10).to_string())

print(f"\nInput Create Time year distribution:")
card_in['yr'] = pd.to_datetime(card_in['Create Time'], errors='coerce').dt.year
print(card_in['yr'].value_counts().sort_index().to_string())

print(f"\nRef Total Days range:")
print(f"  min: {ref_card['Total  Days'].min()}, max: {ref_card['Total  Days'].max()}")
print(f"\nRef last 2 cols (col10, col11) sample:")
last_cols = ref_card.columns[-2:]
print(ref_card[list(last_cols)].head(10).to_string())

# ── FAN-FAIL-R ──────────────────────────────────────────────────────────────
print()
print("=" * 70)
print("FAN-FAIL-R")
print("=" * 70)
fan_in = pd.read_excel(INPUT_FAN)
ref_fan = ref_df(ref_wb, "FAN-FAIL-R", 4)

print(f"Input rows: {len(fan_in)}, unique IPs: {fan_in['IP'].nunique()}")
print(f"Ref rows:   {len(ref_fan)}, unique IPs: {ref_fan['Node IP'].dropna().nunique()}")

print(f"\nInput Circle distribution:")
print(fan_in['Circle'].value_counts().to_string())
print(f"\nRef Circle distribution:")
print(ref_fan['CIRCLE'].value_counts().head(20).to_string())

fan_in['yr'] = pd.to_datetime(fan_in['Create Time'], errors='coerce').dt.year
print(f"\nInput Create Time year distribution:")
print(fan_in['yr'].value_counts().sort_index().to_string())

# ── DEVICE-OFF-R ────────────────────────────────────────────────────────────
print()
print("=" * 70)
print("DEVICE-OFF-R")
print("=" * 70)
dev_in = pd.read_excel(INPUT_DEV)
ref_dev = ref_df(ref_wb, "DEVICE-OFF-R", 4)

print(f"Input rows: {len(dev_in)}, unique IPs: {dev_in['IP'].nunique()}")
print(f"Ref rows:   {len(ref_dev)}, unique IPs: {ref_dev['Node IP'].dropna().nunique()}")

print(f"\nInput CIRCLE distribution:")
print(dev_in['CIRCLE'].value_counts().to_string())
print(f"\nRef CIRCLE distribution:")
print(ref_dev['CIRCLE'].value_counts().head(20).to_string())

# ── DL-FAIL-R ───────────────────────────────────────────────────────────────
print()
print("=" * 70)
print("DL-FAIL-R")
print("=" * 70)
dl_in = pd.read_excel(INPUT_DL)
ref_dl = ref_df(ref_wb, "DL-FAIL-R", 5)

print(f"Input rows: {len(dl_in)}, cols: {list(dl_in.columns)}")
print(f"Ref rows:   {len(ref_dl)}, cols: {list(ref_dl.columns)[:10]}")

print(f"\nInput Circle distribution:")
circ_col = [c for c in dl_in.columns if 'circle' in c.lower() or 'Circle' in c]
if circ_col:
    print(dl_in[circ_col[0]].value_counts().to_string())
print(f"\nRef Circle distribution:")
circ_col2 = [c for c in ref_dl.columns if 'circle' in c.lower() or 'Circle' in c]
if circ_col2:
    print(ref_dl[circ_col2[0]].value_counts().head(20).to_string())

print()
print("=" * 70)
print("MISSING SHEETS in our generated output")
print("=" * 70)
print("1. SUM-TABLE  - summary/pivot table by Circle/BA/SSA")
print("2. DEGR-DL-R  - degraded DL report (what is this?)")
print(f"\nDEGR-DL-R ref rows: {ref_wb['DEGR-DL-R'].max_row}")
sh_degr = ref_wb['DEGR-DL-R']
for r in range(4, 8):
    print(f"  Row {r}: {[str(c.value)[:25] if c.value else '-' for c in sh_degr[r]][:10]}")

print()
print("DASH-DOWN-R comparison:")
ref_dash = ref_df(ref_wb, "DASH-DOWN-R", 5)
print(f"Ref DASH-DOWN-R rows: {len(ref_dash)}, cols: {list(ref_dash.columns)}")
print(ref_dash.head(5).to_string())
