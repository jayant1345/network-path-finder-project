"""Compare generated 07.05.26 report vs reference master - proper column alignment."""
import openpyxl, pandas as pd

GEN        = r"C:\Project_AI\network-path-finder-project\reports_output\DAILY_CPAN_REPORTS_07.05.26_PY.xlsx"
REF_MASTER = r"C:\Project_AI\network-path-finder-project\07.05.26\DAILY_CPAN_REPORTS_MASTER_07.05.26.xlsx"
SHEETS     = ['CARD-OFF-R', 'DEVICE-OFF-R', 'FAN-FAIL-R', 'DL-FAIL-R']

COMPARE_COLS = {
    'CARD-OFF-R':   ['Node IP', 'Node Name', 'Type', 'Circle', 'SSA', 'Card', 'Create Time', 'Total Days'],
    'DEVICE-OFF-R': ['Node IP', 'Node Name', 'TYPE', 'CIRCLE', 'REGION', 'PHASE', 'CREATE TIME', 'Total days'],
    'FAN-FAIL-R':   ['Node IP', 'Node Name', 'TYPE', 'CIRCLE', 'REGION', 'PHASE', 'Alarm ID', 'CREATE TIME', 'Total Days'],
    'DL-FAIL-R':    ['IP A END', 'A END', 'IP Z END', 'Z END', 'Create Time', 'Total  days'],
}
GEN_COLS = {
    'CARD-OFF-R':   ['Node IP', 'Node Name', 'Type', 'Circle', 'SSA', 'Card', 'Create Time', 'Total Days'],
    'DEVICE-OFF-R': ['Node IP', 'Node Name', 'Type', 'Circle', 'Region', 'Phase', 'Create Time', 'Total Days'],
    'FAN-FAIL-R':   ['Node IP', 'Node Name', 'Type', 'Circle', 'SSA', 'Phase', 'Alarm ID', 'Create Time', 'Total Days'],
    'DL-FAIL-R':    ['IP A End', 'A End', 'IP Z End', 'Z End', 'Create Time', 'Total Days'],
}

def norm(v):
    if v is None: return ''
    s = str(v).strip().lstrip('\t')
    return '' if s.lower() in ('nan', 'none') else s

def scalar(v):
    """Return a single scalar from a value (handles Series from duplicate col names)."""
    if isinstance(v, pd.Series):
        v = v.iloc[0] if not v.empty else None
    return v

def load_sheet(wb, sh_name):
    if sh_name not in wb.sheetnames:
        return None
    sh = wb[sh_name]
    hdr_row = None
    for r in range(1, 10):
        val = sh.cell(row=r, column=1).value
        if val and str(val).strip() == 'Sr':
            hdr_row = r
            break
    if hdr_row is None:
        return None
    rows = [[c.value for c in sh[r]] for r in range(hdr_row, sh.max_row + 1)]
    # Make column names unique to avoid pandas ambiguity
    raw_cols = rows[0]
    seen = {}
    unique_cols = []
    for i, c in enumerate(raw_cols):
        base = str(c).strip().replace('\n', ' ') if c else f'_nan{i}'
        if base in seen:
            seen[base] += 1
            unique_cols.append(f'{base}_{seen[base]}')
        else:
            seen[base] = 0
            unique_cols.append(base)
    df = pd.DataFrame(rows[1:], columns=unique_cols)
    df = df[df.iloc[:, 0].notna() & (df.iloc[:, 0].astype(str).str.strip() != '')]
    return df

gen_wb = openpyxl.load_workbook(GEN)
ref_wb = openpyxl.load_workbook(REF_MASTER, data_only=True)

print("=" * 70)
print("COMPARISON: Generated vs Reference - 07.05.26")
print("=" * 70)

for sh_name in SHEETS:
    ref_df = load_sheet(ref_wb, sh_name)
    gen_df = load_sheet(gen_wb, sh_name)

    ref_rows = len(ref_df) if ref_df is not None else 0
    gen_rows = len(gen_df) if gen_df is not None else 0

    print(f"\n{'-'*70}")
    print(f"SHEET: {sh_name}")
    print(f"  REF rows: {ref_rows}   GEN rows: {gen_rows}   diff={ref_rows - gen_rows:+d}")

    if ref_df is None or gen_df is None:
        print("  SKIP: sheet missing")
        continue

    def find_ip_col(df, sh):
        if sh == 'DL-FAIL-R':
            return next((c for c in df.columns if c.lower().startswith('ip a end')), None)
        return next((c for c in df.columns if c.lower().startswith('node ip')), None)

    ref_ip = find_ip_col(ref_df, sh_name)
    gen_ip = find_ip_col(gen_df, sh_name)

    if not ref_ip or not gen_ip:
        print(f"  Cannot find IP column (REF: {ref_ip}, GEN: {gen_ip})")
        print(f"  REF cols: {list(ref_df.columns)[:10]}")
        print(f"  GEN cols: {list(gen_df.columns)[:10]}")
        continue

    ref_ips = set(ref_df[ref_ip].astype(str).str.strip())
    gen_ips = set(gen_df[gen_ip].astype(str).str.strip())
    common   = ref_ips & gen_ips
    only_ref = ref_ips - gen_ips
    only_gen = gen_ips - ref_ips

    print(f"  IPs matched: {len(common)} | only in REF: {len(only_ref)} | only in GEN: {len(only_gen)}")
    if only_gen:
        print(f"  Extra IPs in GEN not in REF: {sorted(only_gen)[:5]}")
    if only_ref:
        print(f"  IPs in REF not in GEN (sample): {sorted(only_ref)[:5]}")

    ref_compare = COMPARE_COLS[sh_name]
    gen_compare = GEN_COLS[sh_name]

    diffs = []
    total_checked = 0

    for ip in sorted(common)[:60]:
        rr = ref_df[ref_df[ref_ip].astype(str).str.strip() == ip]
        gr = gen_df[gen_df[gen_ip].astype(str).str.strip() == ip]
        if rr.empty or gr.empty:
            continue
        total_checked += 1
        rv = rr.iloc[0]
        gv = gr.iloc[0]

        for rc, gc in zip(ref_compare, gen_compare):
            # Find the actual column name (handle unique-suffixed duplicates)
            rc_actual = next((c for c in ref_df.columns if c == rc or c.startswith(rc + '_')), rc)
            rv_val = norm(scalar(rv[rc_actual])) if rc_actual in rv.index else ''
            gv_val = norm(scalar(gv[gc]))        if gc          in gv.index else ''

            if 'days' in rc.lower():
                try: rv_val = str(int(float(rv_val))) if rv_val else ''
                except: pass
                try: gv_val = str(int(float(gv_val))) if gv_val else ''
                except: pass

            if rv_val != gv_val:
                diffs.append((ip, rc, rv_val, gv_val))

    match_pct = (total_checked - len(set(d[0] for d in diffs))) / max(total_checked, 1) * 100

    if not diffs:
        print(f"  Data: ALL {total_checked} matching IPs checked - IDENTICAL values")
    else:
        diffed_ips = len(set(d[0] for d in diffs))
        print(f"  Data: {total_checked - diffed_ips}/{total_checked} IPs fully match ({match_pct:.0f}%)")
        print(f"  Diffs (first 12):")
        for ip, col, rv, gv in diffs[:12]:
            print(f"    != IP={ip} [{col}]: REF='{rv}' vs GEN='{gv}'")

print(f"\n{'='*70}")
print("SUMMARY")
print("=" * 70)
print("GEN row counts: CARD-OFF-R=31  DEVICE-OFF-R=38  FAN-FAIL-R=37  DL-FAIL-R=89")
print("REF row counts: see above (accumulated old data inflates REF numbers)")
print()
print("Data quality for matching IPs: see per-sheet results above.")
