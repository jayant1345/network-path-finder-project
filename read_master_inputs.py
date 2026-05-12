"""Read the actual input data pasted INTO the master template for 07.05.26."""
import openpyxl, pandas as pd

MASTER = r"C:\Project_AI\network-path-finder-project\07.05.26\DAILY_CPAN_REPORTS_MASTER_07.05.26.xlsx"

print("Opening master template (read_only, no formulas)...")
wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)

def read_input_sheet(wb, sh_name, max_rows=2000):
    """Read raw pasted data from an input sheet."""
    if sh_name not in wb.sheetnames:
        return None
    sh = wb[sh_name]
    rows = []
    header = None
    for i, r in enumerate(sh.iter_rows(max_row=max_rows, values_only=True)):
        if not any(v for v in r if v is not None):
            if header is not None:
                break   # stop on first all-empty row after data
            continue
        if header is None:
            header = [str(v).strip() if v else f'col{i}' for i, v in enumerate(r)]
        else:
            rows.append(r)
    if header is None or not rows:
        return None
    df = pd.DataFrame(rows, columns=header)
    return df

for sh_name in ['CARD-OFF', 'DEVICE-OFF', 'FAN-FAIL', 'DL-FAIL', 'DL-EMS']:
    df = read_input_sheet(wb, sh_name)
    if df is None:
        print(f"\n{sh_name}: empty or not found")
        continue
    print(f"\n{'='*60}")
    print(f"{sh_name}: {len(df)} rows | cols: {list(df.columns)[:8]}")
    # Circle distribution
    circ_col = next((c for c in df.columns if 'circle' in c.lower() or c=='CIRCLE'), None)
    if circ_col:
        print(f"  CIRCLE distribution: {df[circ_col].value_counts().to_string()}")
    print(f"  Sample row 1: {df.iloc[0].tolist()[:8]}" if len(df)>0 else "  (empty)")
