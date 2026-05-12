"""Test report_gen with 02.04.26 input files and compare against reference."""
import sys, os
sys.path.insert(0, r"C:\Project_AI\network-path-finder-project")
from datetime import datetime
from report_gen import generate_report, read_card_off, read_device_off, read_fan_fail, read_dl_fail

REPORT_DATE = datetime(2026, 4, 2)
DATE_STR    = "02.04.26"
OUT         = r"C:\Project_AI\network-path-finder-project\reports_output\TEST_NEW_02.04.26.xlsx"

INPUT_CARD = r"C:\Project_AI\cpan_report\CARD OFFLINE 02-04-2026.xlsx"
INPUT_DEV  = r"C:\Project_AI\cpan_report\DEVICE OFFLINE 02-04-2026.xlsx"
INPUT_FAN  = r"C:\Project_AI\cpan_report\FAN FAILURE 02-04-2026.xlsx"
INPUT_DL   = r"C:\Project_AI\cpan_report\CPAN DL FAIL REPORT  02-04-2026.xlsx"

# Reference counts (from reference report DAILY_CPAN_REPORTS_02.04.26.xlsx)
REF_COUNTS = {
    'CARD-OFF-R':   200,
    'DEVICE-OFF-R': 150,
    'FAN-FAIL-R':   151,
    'DL-FAIL-R':    309,
}

print("=" * 60)
print(f"Test: 02.04.26 | Circles: GJ + WTR")
print("=" * 60)

# Individual sheet tests
print("\nCARD-OFF-R:")
df = read_card_off(INPUT_CARD, REPORT_DATE)
print(f"  Generated: {len(df)} rows  (Reference: {REF_COUNTS['CARD-OFF-R']})")
print(f"  Circles in output: {df['Circle'].unique().tolist() if 'Circle' in df.columns else 'N/A'}")
print(f"  Sample Alarm IDs: N/A (CARD has no Alarm ID)")
print(f"  Sample row 1: {df.iloc[0].to_dict() if len(df) > 0 else 'EMPTY'}")
print(f"  Year-2000 rows remaining: {(df['Total Days'].astype(str).str.contains('9587') if 'Total Days' in df.columns else False).any()}")

print("\nDEVICE-OFF-R:")
df2 = read_device_off(INPUT_DEV, REPORT_DATE)
print(f"  Generated: {len(df2)} rows  (Reference: {REF_COUNTS['DEVICE-OFF-R']})")
print(f"  Circles: {df2['Circle'].unique().tolist() if 'Circle' in df2.columns else 'N/A'}")

print("\nFAN-FAIL-R:")
df3 = read_fan_fail(INPUT_FAN, REPORT_DATE)
print(f"  Generated: {len(df3)} rows  (Reference: {REF_COUNTS['FAN-FAIL-R']})")
print(f"  Circles: {df3['Circle'].unique().tolist() if 'Circle' in df3.columns else 'N/A'}")
if 'Alarm ID' in df3.columns and len(df3):
    sample_ids = df3['Alarm ID'].head(5).tolist()
    print(f"  Sample Alarm IDs (should have no tab): {sample_ids}")

print("\nDL-FAIL-R:")
df4 = read_dl_fail(INPUT_DL, REPORT_DATE)
print(f"  Generated: {len(df4)} rows  (Reference: {REF_COUNTS['DL-FAIL-R']})")
print(f"  Circles: {df4['Circle'].unique().tolist() if 'Circle' in df4.columns else 'N/A'}")

# Generate full report
print("\n" + "=" * 60)
print("Generating full report...")
file_map = {
    'card_off':   INPUT_CARD,
    'device_off': INPUT_DEV,
    'fan_fail':   INPUT_FAN,
    'dl_fail':    INPUT_DL,
    'dash_down':  None,
}
logs = generate_report(file_map, REPORT_DATE, OUT)
for log in logs:
    print(log)

print("\nComparison vs reference:")
import openpyxl
gen_wb = openpyxl.load_workbook(OUT)
print(f"  Sheets: {gen_wb.sheetnames}")
for sh_name, ref_count in REF_COUNTS.items():
    if sh_name in gen_wb.sheetnames:
        sh = gen_wb[sh_name]
        gen_count = sh.max_row - 2   # subtract title + header rows
        match = "OK" if abs(gen_count - ref_count) <= 2 else "MISMATCH"
        print(f"  {sh_name}: gen={gen_count}, ref={ref_count}  [{match}]")
    else:
        print(f"  {sh_name}: MISSING")
