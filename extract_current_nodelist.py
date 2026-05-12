"""Extract current NODE-LIST and DL-EMS IPs from 07.05.26 master template."""
import openpyxl, pandas as pd, re

MASTER  = r"C:\Project_AI\network-path-finder-project\07.05.26\DAILY_CPAN_REPORTS_MASTER_07.05.26.xlsx"
OUT_NL  = r"C:\Project_AI\network-path-finder-project\node_list.csv"
OUT_DL  = r"C:\Project_AI\network-path-finder-project\dl_ems_ips.csv"

IP_RE = re.compile(r'\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b')

print("Opening 07.05.26 master (full mode)...")
wb = openpyxl.load_workbook(MASTER, data_only=True)

# â”€â”€ NODE-LIST â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print("\n=== Extracting NODE-LIST ===")
sh = wb['NODE-LIST']
node_ips = set()
hdr = None
for r in range(1, sh.max_row + 1):
    row = [c.value for c in sh[r]]
    if not any(v for v in row if v is not None):
        continue
    if hdr is None:
        hdr = [str(v).strip().replace('\n',' ') if v else f'c{i}'
               for i, v in enumerate(row)]
        continue
    # Extract Node IP column
    ip_col_idx = next((i for i, h in enumerate(hdr) if 'node ip' in h.lower() or 'ip' in h.lower()), None)
    if ip_col_idx is not None and ip_col_idx < len(row):
        val = str(row[ip_col_idx]).strip() if row[ip_col_idx] else ''
        m = IP_RE.search(val)
        if m:
            node_ips.add(m.group(1))
    # Also try to find IP anywhere in the 'Node Name' column (format: IP_NAME_...)
    name_col_idx = next((i for i, h in enumerate(hdr) if 'node name' in h.lower()), None)
    if name_col_idx is not None and name_col_idx < len(row):
        val = str(row[name_col_idx]).strip() if row[name_col_idx] else ''
        m = IP_RE.match(val)
        if m:
            node_ips.add(m.group(1))

print(f"  NODE-LIST unique IPs: {len(node_ips)}")
# Also extract IPs from CARD-OFF master input (cols 2,3 show SSA/Circle; col 6 = IP)
print("\n=== Checking SSACODE for SSAâ†’Circle mapping ===")
sh_ssacode = wb['SSACODE']
ssa_gj = set()
ssa_wtr_gj = set()
for r in sh_ssacode.iter_rows(min_row=3, values_only=True):
    circle = str(r[4]).strip() if r[4] else ''
    ssa    = str(r[5]).strip() if r[5] else ''
    if circle == 'GJ' and ssa:
        ssa_gj.add(ssa)
    elif circle == 'WTR' and ssa:
        ssa_wtr_gj.add(ssa)
print(f"  GJ SSAs: {sorted(ssa_gj)}")
print(f"  WTR-GJ SSAs: {sorted(ssa_wtr_gj)}")

# â”€â”€ DL-EMS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print("\n=== Extracting DL-EMS IPs ===")
sh_dlems = wb['DL-EMS']
dl_ips = set()
for r in range(10, sh_dlems.max_row + 1):
    row = [c.value for c in sh_dlems[r]]
    # Col 4 (index 4) has the DL Name with IPs: 'SSA AM-10.121.18.249'
    if len(row) > 4 and row[4]:
        val = str(row[4])
        for m in IP_RE.finditer(val):
            dl_ips.add(m.group(1))
    # Col 5 has the raw name with tabs: '\tSSA AM-10.121.18.24...'
    if len(row) > 5 and row[5]:
        val = str(row[5])
        for m in IP_RE.finditer(val):
            dl_ips.add(m.group(1))

print(f"  DL-EMS unique IPs: {len(dl_ips)}")
print(f"  Sample DL-EMS IPs: {list(dl_ips)[:10]}")

# Save
nl_df = pd.DataFrame(sorted(node_ips), columns=['Node IP'])
nl_df.to_csv(OUT_NL, index=False)
print(f"\nSaved node_list.csv: {len(nl_df)} IPs â†’ {OUT_NL}")

dl_df = pd.DataFrame(sorted(dl_ips), columns=['Node IP'])
dl_df.to_csv(OUT_DL, index=False)
print(f"Saved dl_ems_ips.csv: {len(dl_df)} IPs â†’ {OUT_DL}")

# â”€â”€ Quick verification â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
print("\n=== Verification: CARD-OFF filter using new NODE-LIST ===")
card_in = pd.read_excel(
    r"C:\Project_AI\network-path-finder-project\07.05.26\CARD OFFLINE 07.05.2026.xlsx",
    dtype=str)
card_in.columns = [c.strip() for c in card_in.columns]
ip_col = 'IP'
after_nl = card_in[card_in[ip_col].astype(str).str.strip().isin(node_ips)]
print(f"  Input: {len(card_in)} rows â†’ after NODE-LIST filter: {len(after_nl)} [REF=200]")
print(f"  Circle dist in filtered: {after_nl['CIRCLE'].value_counts().to_string()}")

print("\n=== Verification: DL-FAIL filter using DL-EMS IPs ===")
dl_in = pd.read_excel(
    r"C:\Project_AI\network-path-finder-project\07.05.26\DL FAIL REPORT 07.05.2026.xlsx",
    dtype=str)
dl_in.columns = [c.strip() for c in dl_in.columns]
# Rename A/Z end columns
dl_in = dl_in.rename(columns={'IP A END': 'IP_A', 'IP Z END': 'IP_Z'})
if 'IP_A' in dl_in.columns and 'IP_Z' in dl_in.columns:
    a_in = dl_in['IP_A'].astype(str).str.strip().isin(dl_ips)
    z_in = dl_in['IP_Z'].astype(str).str.strip().isin(dl_ips)
    after_dl = dl_in[a_in | z_in]
    print(f"  Input: {len(dl_in)} rows â†’ after DL-EMS filter: {len(after_dl)} [REF=300]")

