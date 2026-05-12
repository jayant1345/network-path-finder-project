"""Extract SSAâ†’BA mapping from reference report and SUM-TABLE."""
import openpyxl
import pandas as pd

REF = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_02.04.26.xlsx"
MASTER = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_MASTER_02.04.26.xlsx"

ref_wb = openpyxl.load_workbook(REF, data_only=True)

# Extract from SUM-TABLE - columns: Circle, BA, SSA, ...
print("=== SUM-TABLE BA-SSA mapping ===")
sh = ref_wb['SUM-TABLE']
ba_map = {}
for r in range(6, sh.max_row + 1):
    row = [c.value for c in sh[r]]
    circle = str(row[0]).strip() if row[0] else None
    ba = str(row[1]).strip() if row[1] else None
    ssa = str(row[2]).strip() if row[2] else None
    if ssa and ba and circle and ssa not in ('None', 'nan'):
        ba_map[ssa] = (circle, ba)

# Print sorted by circle then ba then ssa
for ssa in sorted(ba_map, key=lambda x: (ba_map[x][0], ba_map[x][1], x)):
    circle, ba = ba_map[ssa]
    print(f"  {circle:6s}  {ba:20s}  {ssa}")

print(f"\nTotal SSA entries: {len(ba_map)}")

# Also extract from CARD-OFF-R data rows (has BA in col 10)
print("\n=== CARD-OFF-R SSA-BA from data rows ===")
sh_card = ref_wb['CARD-OFF-R']
card_ba = {}
for r in range(5, sh_card.max_row + 1):
    row = [c.value for c in sh_card[r]]
    if len(row) >= 11:
        ssa = str(row[5]).strip() if row[5] else None
        ba_val = str(row[9]).strip() if len(row) > 9 and row[9] else None
        if ssa and ba_val and ssa not in ('None', 'nan'):
            card_ba[ssa] = ba_val

for ssa in sorted(card_ba):
    print(f"  {ssa:20s} â†’ {card_ba[ssa]}")

# Check if master file has a mapping sheet
print("\n=== MASTER FILE sheets ===")
try:
    master_wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)
    print(f"Sheets: {master_wb.sheetnames}")
    # Look for a sheet with circle/BA/SSA mapping
    for sh_name in master_wb.sheetnames:
        sh = master_wb[sh_name]
        row1 = [str(c.value)[:20] if c.value else '-' for c in next(sh.iter_rows(max_row=1))]
        print(f"  {sh_name}: {row1[:8]}")
except Exception as e:
    print(f"  Error reading master: {e}")

