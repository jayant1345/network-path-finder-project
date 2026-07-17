"""
Scrape GCS (MantisBT) dockets for a shift day (08:00 -> next day 08:00) and
write the organization's report: Reporter=99412764 excluded from Dockets/CPAN;
their CPAN-DL dockets reinstated into Other, MAAN-TL reinstated into MAAN;
CPAN sheet is clean minor-severity, non-99412764 dockets only.

Used by server.py's /docket_report/* routes (background job pattern).
"""
import os
import re
import urllib.request
from datetime import datetime, timedelta

import pandas as pd
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

load_dotenv()

BASE_URL = "https://intranetguj.bsnl.co.in/gcs"

COLUMNS = [
    "Docket ID", "Project", "SSA_NAME", "Category", "Work Type", "Amount of Work",
    "Severity", "Reporter (Booked By)", "Status", "Summary",
    "Date Submitted", "Cleared By", "Cleared On", "Time to Clear",
]
TEXT_COLUMNS = {"Docket ID", "Reporter (Booked By)", "Cleared By"}
EXCLUDE_REPORTER = "99412764"

HEADER_COLORS = {
    "Final Report": "8B0000",
    "Dockets": "1F4E78",
    "CPAN": "2E7D32",
    "MAAN": "6A1B9A",
    "Other": "B45309",
    "Summary": "00695C",
}
SEVERITY_FILL_COLORS = {"major": "FFC7CE", "minor": "C6EFCE"}

# Colors mirroring the DAILY_CPAN_DOCKET_CALCULATION reference workbook's
# FINAL-TABLE sheet: light-cyan title bar, cyan headers, light-green data
# rows, magenta "Returned in Feedback" text.
FINAL_TITLE_BG = "B2EBF2"
FINAL_HEADER_BG = "00E5E5"
FINAL_ROW_BG = "C6EFCE"
FINAL_FEEDBACK_TEXT = "FF00FF"

# CPAN shift schedule: 08:00-15:30, 12:30-20:30, 20:00-08:30 (next day). The
# reporting day runs 08:00 -> next day 08:00, and the day/night duty split
# used by the reference workbook (DAILY_CPAN_DOCKET_CALCULATION) is 20:00 (8 PM).
SHIFT_START_HOUR = 8
NIGHT_START_HOUR = 20

CPAN_STAFF_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cpan_staff_ids.txt")


def _load_cpan_staff_roster():
    """Reads CPAN_STAFF_FILE (one staff ID per line, '#' starts a comment,
    blank lines ignored) if it exists and is non-empty. Returns a set of IDs,
    or None if not configured yet - callers must treat None as "no filtering,
    count everyone" so the report behaves exactly as before this feature
    existed until the roster is actually provided."""
    if not os.path.isfile(CPAN_STAFF_FILE):
        return None
    ids = set()
    with open(CPAN_STAFF_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.split("#", 1)[0].strip()
            if line:
                ids.add(line)
    return ids or None


def shift_window(date_str):
    """Returns (start_dt, end_dt) for the shift day starting 08:00 on date_str
    (DD-MM-YYYY) through 08:00 the next day."""
    day, month, year = (int(x) for x in date_str.split("-"))
    start_dt = datetime(year, month, day, SHIFT_START_HOUR, 0, 0)
    end_dt = start_dt + timedelta(days=1)
    return start_dt, end_dt


def _login(page, username, password, login_url):
    page.goto(login_url, wait_until="networkidle")
    page.fill("#username", username)
    page.click("form#login-form input[type=submit]")
    page.wait_for_load_state("networkidle")
    page.fill("#password", password)
    page.click("form#login-form input[type=submit]")
    page.wait_for_load_state("networkidle")


def _build_search_url(start_dt, end_dt):
    """Searches by Date Submitted across both calendar dates spanned by the
    shift window (the site's filter only has day granularity); exact-timestamp
    filtering to the real 08:00-08:00 window happens client-side afterward."""
    return (
        f"{BASE_URL}/search.php"
        "?project_id=0&sticky=on&sort=id&dir=ASC"
        "&hide_status=-2&status=0"
        "&filter_by_date=on"
        f"&start_day={start_dt.day}&end_day={end_dt.day}"
        f"&start_month={start_dt.month}&end_month={end_dt.month}"
        f"&start_year={start_dt.year}&end_year={end_dt.year}"
        "&match_type=0&per_page=1000"
    )


def _build_last_updated_search_url(start_dt, end_dt):
    """Searches by Last Updated instead of Date Submitted — catches dockets
    booked long before this shift window that were finally resolved/touched
    during it (MantisBT has no OR-across-filter-types search, so this is a
    separate query merged with the submitted-date one in Python)."""
    return (
        f"{BASE_URL}/search.php"
        "?project_id=0&sticky=on&sort=id&dir=ASC"
        "&hide_status=-2&status=0"
        "&filter_by_last_updated_date=on"
        f"&last_updated_start_day={start_dt.day}&last_updated_end_day={end_dt.day}"
        f"&last_updated_start_month={start_dt.month}&last_updated_end_month={end_dt.month}"
        f"&last_updated_start_year={start_dt.year}&last_updated_end_year={end_dt.year}"
        "&match_type=0&per_page=1000"
    )


def _parse_bug_list(html):
    soup = BeautifulSoup(html, "html.parser")
    rows = soup.select("table#buglist tbody tr")
    issues = []
    for r in rows:
        id_link = r.select_one("td.column-id a")
        if not id_link:
            continue
        reporter_cell = r.select_one("td.column-reporter")
        reporter_link = reporter_cell.find("a")
        issues.append(
            {
                "Docket ID": id_link.get_text(strip=True),
                "Project": r.select_one("td.column-project-id").get_text(strip=True),
                "SSA_NAME": r.select_one("td.column-custom-SSA_NAME").get_text(strip=True),
                "Category": r.select_one("td.column-category").get_text(strip=True),
                "Severity": r.select_one("td.column-severity").get_text(strip=True),
                "Reporter (Booked By)": reporter_cell.get_text(strip=True),
                "Reporter Name": reporter_link.get("title", "") if reporter_link else "",
                "Status": r.select_one("td.column-status").get_text(" ", strip=True),
                "Summary": r.select_one("td.column-summary").get_text(strip=True),
            }
        )
    return issues


def _find_header_value(soup, header_text):
    """Find a <tr> whose cells contain header_text, return the matching value
    from the next sibling row at the same cell index."""
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        texts = [c.get_text(strip=True) for c in cells]
        if header_text in texts:
            idx = texts.index(header_text)
            value_tr = tr.find_next_sibling("tr")
            if value_tr:
                value_cells = value_tr.find_all(["td", "th"])
                if idx < len(value_cells):
                    return value_cells[idx].get_text(strip=True)
    return ""


def _find_inline_value(soup, label):
    """Find a 2-cell <tr> like ['Amount_of_work', '5'] and return the value cell."""
    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        if len(cells) == 2 and cells[0].get_text(strip=True) == label:
            return cells[1].get_text(strip=True)
    return ""


def _parse_issue_detail(html):
    soup = BeautifulSoup(html, "html.parser")
    date_submitted = _find_header_value(soup, "Date Submitted")
    amount_of_work = _find_inline_value(soup, "Amount_of_work")

    cleared_on, cleared_by, cleared_by_name = "", "", ""
    history_table = None
    for table in soup.find_all("table"):
        rows = table.select("tr")
        first_row_texts = [c.get_text(strip=True) for c in rows[0].find_all(["td", "th"])] if rows else []
        if first_row_texts == ["Date Modified", "Username", "Field", "Change"]:
            history_table = table
            break

    if history_table:
        for r in history_table.select("tr")[1:]:  # skip header row
            cell_els = r.find_all(["td", "th"])
            if len(cell_els) != 4:
                continue
            date_modified, username, field, change = [c.get_text(strip=True) for c in cell_els]
            # Only "-> resolved" credits the technical fix. "-> closed" is a
            # separate, later confirmation step - often done by the reporter
            # (field staff) rather than whoever actually did the work - so it
            # must NOT overwrite cleared_by. If a docket is reopened and
            # resolved again, the loop keeps scanning and the last "->
            # resolved" wins, correctly reflecting the final fix.
            if field == "Status" and re.search(r"=>\s*resolved", change, re.I):
                cleared_on, cleared_by = date_modified, username
                username_link = cell_els[1].find("a")
                cleared_by_name = username_link.get("title", "") if username_link else ""

    return date_submitted, cleared_on, cleared_by, cleared_by_name, amount_of_work


def _strip_category_prefix(raw_category):
    """Strip the hidden '[PROJECT]' prefix GCS embeds in the category text."""
    return re.sub(r"^\[[^\]]*\]", "", raw_category).strip()


def _normalize_category(clean_category):
    """Fold 'General' (a data-entry catch-all for maintenance work) into Complaint."""
    return "Complaint" if clean_category.strip().lower() == "general" else clean_category.strip()


def _classify_work_type(normalized_category):
    """Service = Service Creation docket; everything else (Complaint/General) = Maintenance."""
    return "Service Creation" if normalized_category.strip().lower() == "service" else "Maintenance"


def _normalize_amount_of_work(raw):
    """Amount of work of 0 or blank still represents 1 unit of completed work."""
    try:
        val = int(raw)
    except (TypeError, ValueError):
        val = 0
    return 1 if val == 0 else val


def _compute_duration(start_str, end_str):
    if not start_str or not end_str:
        return ""
    fmt = "%Y-%m-%d %H:%M"
    delta = datetime.strptime(end_str, fmt) - datetime.strptime(start_str, fmt)
    total_minutes = int(delta.total_seconds() // 60)
    days, rem_minutes = divmod(total_minutes, 24 * 60)
    hours, minutes = divmod(rem_minutes, 60)
    if days:
        return f"{days}d {hours:02d}h {minutes:02d}m"
    return f"{hours:02d}h {minutes:02d}m"


def _format_duration_minutes(total_minutes):
    """Like _compute_duration but from a float minute count (e.g. an average)."""
    if total_minutes is None or pd.isna(total_minutes):
        return ""
    total_minutes = int(round(total_minutes))
    days, rem_minutes = divmod(total_minutes, 24 * 60)
    hours, minutes = divmod(rem_minutes, 60)
    if days:
        return f"{days}d {hours:02d}h {minutes:02d}m"
    return f"{hours:02d}h {minutes:02d}m"


def _avg_clear_minutes(df, work_type=None):
    """Average minutes from Date Submitted to Cleared On, optionally filtered
    to one Work Type. df should already be restricted to cleared dockets."""
    if work_type:
        df = df[df["Work Type"] == work_type]
    if df.empty:
        return None
    submitted = pd.to_datetime(df["Date Submitted"], format="%Y-%m-%d %H:%M", errors="coerce")
    cleared = pd.to_datetime(df["Cleared On"], format="%Y-%m-%d %H:%M", errors="coerce")
    minutes = (cleared - submitted).dt.total_seconds() / 60
    return minutes.mean()


def _avg_clear_time_by_group(df, group_col):
    """Per value of group_col: average clear time (formatted string) for
    Service Creation and Maintenance dockets. df should be cleared dockets only."""
    result = {}
    for group_val, sub in df.groupby(group_col):
        result[group_val] = (
            _format_duration_minutes(_avg_clear_minutes(sub, "Service Creation")),
            _format_duration_minutes(_avg_clear_minutes(sub, "Maintenance")),
        )
    return result


def _detect_proxy(log=print):
    """Playwright's Chromium does not pick up Windows' configured LAN proxy on
    its own (unlike a normal browser), so on a network where GCS is only
    reachable via a proxy - as opposed to a direct WiFi connection - the
    browser fails with ERR_NAME_NOT_RESOLVED unless we hand it the proxy
    explicitly. Set GCS_PROXY_SERVER in .env (e.g. http://10.1.1.1:8080) to
    override auto-detection if it ever picks the wrong address."""
    override = os.environ.get("GCS_PROXY_SERVER")
    if override:
        log(f"Using proxy from GCS_PROXY_SERVER: {override}")
        return {"server": override}

    server = (urllib.request.getproxies().get("https")
              or urllib.request.getproxies().get("http"))
    if not server:
        return None
    if "://" not in server:
        server = "http://" + server
    log(f"Detected LAN proxy, routing GCS traffic via {server}")
    return {"server": server}


def _enrich_issue(context, issue):
    """Fetches a docket's detail page and fills in its clear/category/work-type
    fields in place (mutates issue)."""
    bug_id = issue["Docket ID"].lstrip("0") or "0"
    resp = context.request.get(f"{BASE_URL}/view.php?id={bug_id}")
    date_submitted, cleared_on, cleared_by, cleared_by_name, amount_of_work = _parse_issue_detail(resp.text())
    raw_category = _strip_category_prefix(issue["Category"])  # Complaint/Service/General, pre-fold
    issue["Category (Raw)"] = raw_category
    issue["Category"] = _normalize_category(raw_category)
    issue["Work Type"] = _classify_work_type(issue["Category"])
    issue["Amount of Work"] = _normalize_amount_of_work(amount_of_work)
    issue["Date Submitted"] = date_submitted
    issue["Cleared By"] = cleared_by
    issue["Cleared By Name"] = cleared_by_name
    issue["Cleared On"] = cleared_on
    issue["Time to Clear"] = _compute_duration(date_submitted, cleared_on)


def _scrape_dockets(date_str, username, password, login_url, log=print):
    """date_str: DD-MM-YYYY, the shift day's start date. Returns dockets that
    were either BOOKED in [date_str 08:00, next day 08:00), or CLEARED in that
    window regardless of how long ago they were booked (so old backlog that
    finally gets resolved during this shift is credited to it). Adds two bool
    columns: "Booked This Window" and "Cleared This Window"."""
    start_dt, end_dt = shift_window(date_str)

    with sync_playwright() as p:
        proxy_config = _detect_proxy(log=log)
        browser = p.chromium.launch(headless=True, proxy=proxy_config)
        context = browser.new_context()
        page = context.new_page()

        log("Logging in to GCS...")
        _login(page, username, password, login_url)

        log(f"Searching dockets submitted {start_dt:%d-%m-%Y %H:%M} to {end_dt:%d-%m-%Y %H:%M} ...")
        page.goto(_build_search_url(start_dt, end_dt), wait_until="networkidle")
        by_submitted = _parse_bug_list(page.content())

        log(f"Searching dockets last-updated {start_dt:%d-%m-%Y %H:%M} to {end_dt:%d-%m-%Y %H:%M} "
            f"(catches old backlog cleared during this shift) ...")
        page.goto(_build_last_updated_search_url(start_dt, end_dt), wait_until="networkidle")
        by_updated = _parse_bug_list(page.content())

        issues_by_id = {i["Docket ID"]: i for i in by_submitted}
        for i in by_updated:
            issues_by_id.setdefault(i["Docket ID"], i)
        issues = list(issues_by_id.values())
        log(f"Found {len(by_submitted)} by submit-date + {len(by_updated)} by last-updated "
            f"= {len(issues)} unique candidates. Fetching per-docket clear details...")

        for i, issue in enumerate(issues, 1):
            _enrich_issue(context, issue)
            if i % 20 == 0 or i == len(issues):
                log(f"  {i}/{len(issues)} processed")

        browser.close()

    df = pd.DataFrame(issues)
    submitted_dt = pd.to_datetime(df["Date Submitted"], format="%Y-%m-%d %H:%M", errors="coerce")
    cleared_dt = pd.to_datetime(df["Cleared On"], format="%Y-%m-%d %H:%M", errors="coerce")
    df["Booked This Window"] = (submitted_dt >= start_dt) & (submitted_dt < end_dt)
    df["Cleared This Window"] = (df["Cleared By"] != "") & (cleared_dt >= start_dt) & (cleared_dt < end_dt)
    df = df[df["Booked This Window"] | df["Cleared This Window"]].reset_index(drop=True)

    backlog_cleared = int((df["Cleared This Window"] & ~df["Booked This Window"]).sum())
    log(f"{len(df)} dockets in final report ({int(df['Booked This Window'].sum())} booked this window, "
        f"{backlog_cleared} older backlog cleared during this window).")
    return df


def _keyword_mask(df, keywords):
    mask = pd.Series(True, index=df.index)
    for kw in keywords:
        mask &= df["Summary"].str.contains(re.escape(kw), case=False, na=False)
    return mask


def _pivot_summary(df, group_col, value_col):
    """Count of each value_col category per value of group_col, sorted by total
    descending. Returns (table, value_columns) — value_columns are whatever
    distinct values value_col actually has (so callers don't need to know them)."""
    if df.empty:
        return pd.DataFrame(columns=[group_col, "Total"]), []
    grouped = df.groupby([group_col, value_col]).size().unstack(fill_value=0)
    value_cols = list(grouped.columns)
    grouped["Total"] = grouped[value_cols].sum(axis=1)
    grouped = grouped.reset_index().sort_values("Total", ascending=False).reset_index(drop=True)
    return grouped, value_cols


def _pivot_summary_with_work(df, group_col, value_col, work_col="Amount of Work"):
    """Like _pivot_summary, but adds summed Amount of Work per category too
    (columns suffixed ' (AoW)'). Returns (table, count_cols, aow_cols)."""
    table, value_cols = _pivot_summary(df, group_col, value_col)
    aow_cols = [f"{c} (AoW)" for c in value_cols] + ["Total (AoW)"]

    if df.empty:
        for c in aow_cols:
            table[c] = []
        return table, value_cols, aow_cols

    work_sum = df.groupby([group_col, value_col])[work_col].sum().unstack(fill_value=0)
    for col in value_cols:
        if col not in work_sum.columns:
            work_sum[col] = 0
    work_sum = work_sum[value_cols]
    work_sum["Total (AoW)"] = work_sum[value_cols].sum(axis=1)
    work_sum = work_sum.rename(columns={c: f"{c} (AoW)" for c in value_cols}).reset_index()

    merged = table.merge(work_sum, on=group_col, how="left")
    for c in aow_cols:
        merged[c] = merged[c].fillna(0).astype(int)
    return merged, value_cols, aow_cols


def _name_lookup(df, id_col, name_col):
    named = df[df[name_col] != ""].drop_duplicates(id_col)
    return dict(zip(named[id_col], named[name_col]))


def _write_summary_table(ws, header_fill, start_row, title, id_label, table, value_cols, name_map=None, extra_cols=None, text_cols=None):
    bold = Font(bold=True)
    text_cols = text_cols or []
    all_value_cols = value_cols + ["Total"] + (extra_cols or []) + text_cols
    headers = [id_label] + all_value_cols
    ws.cell(row=start_row, column=1, value=title).font = bold
    header_row = start_row + 1
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for i, (_, r) in enumerate(table.iterrows()):
        row = header_row + 1 + i
        id_value = str(r[id_label])
        id_cell = ws.cell(row=row, column=1, value=id_value)
        id_cell.number_format = "@"  # preserve leading zeros in usernames
        if name_map and name_map.get(id_value):
            id_cell.comment = Comment(name_map[id_value], "GCS")
        for j, col in enumerate(all_value_cols, 2):
            ws.cell(row=row, column=j, value=r[col] if col in text_cols else int(r[col]))
    return header_row + 1 + len(table) + 2  # next free row, one blank line gap


def _write_summary_sheet(writer, df, cpan_roster=None):
    """Writes the Summary sheet and returns (ws, next_free_row) so callers can
    append additional tables (e.g. _write_report_v2's CPAN staff summary).

    "Booked by" tables are scoped to dockets booked in this shift window.
    "Cleared by Person" is scoped to dockets cleared in this window, which can
    include old backlog booked earlier — that work is credited here.

    cpan_roster: if given, a docket cleared by someone not in the roster is
    treated as not resolved (falls into "pending" here too), so this sheet's
    "Cleared This Window" / "Pending" figures stay consistent with the Final
    Report's roster-filtered "Total Resolved" / "Total Pending Dockets"."""
    is_cleared = df["Cleared This Window"]
    if cpan_roster is not None:
        is_cleared = is_cleared & df["Cleared By"].isin(cpan_roster)
    booked_df = df[df["Booked This Window"]]
    cleared_df = df[is_cleared]
    pending_df = df[df["Booked This Window"] & ~is_cleared]
    total_count = len(booked_df)
    cleared_count = len(cleared_df)
    pending_count = len(pending_df)

    ssa_table, ssa_cats = _pivot_summary(booked_df, "SSA_NAME", "Category (Raw)")
    reporter_table, wt_cols = _pivot_summary(booked_df, "Reporter (Booked By)", "Work Type")
    cleared_table, wt_cols2, cleared_aow = _pivot_summary_with_work(cleared_df, "Cleared By", "Work Type")
    pending_table, wt_cols3, pending_aow = _pivot_summary_with_work(pending_df, "SSA_NAME", "Work Type")

    # Per-staff KPI: average clear time (Date Submitted -> Cleared On) by category.
    avg_times = _avg_clear_time_by_group(cleared_df, "Cleared By")
    avg_time_cols = ["Avg Time (Creation)", "Avg Time (Maintenance)"]
    cleared_table[avg_time_cols[0]] = cleared_table["Cleared By"].map(lambda x: avg_times.get(x, ("", ""))[0])
    cleared_table[avg_time_cols[1]] = cleared_table["Cleared By"].map(lambda x: avg_times.get(x, ("", ""))[1])

    reporter_names = _name_lookup(df, "Reporter (Booked By)", "Reporter Name")
    cleared_names = _name_lookup(df, "Cleared By", "Cleared By Name")

    ws = writer.book.create_sheet("Summary")
    bold = Font(bold=True)
    header_color = HEADER_COLORS["Summary"]
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    ws.sheet_properties.tabColor = header_color
    max_value_cols = max(len(ssa_cats), len(wt_cols), len(wt_cols2) + len(cleared_aow) + len(avg_time_cols), len(wt_cols3) + len(pending_aow))

    ws.cell(row=1, column=1, value="Total Booked This Window").font = bold
    ws.cell(row=1, column=2, value=total_count)
    ws.cell(row=2, column=1, value="Cleared This Window (incl. old backlog)").font = bold
    ws.cell(row=2, column=2, value=cleared_count)
    ws.cell(row=3, column=1, value="Pending (from this window's bookings)").font = bold
    ws.cell(row=3, column=2, value=pending_count)

    row = _write_summary_table(ws, header_fill, 5, "Dockets Booked by SSA (by Category)", "SSA_NAME", ssa_table, ssa_cats)
    row = _write_summary_table(ws, header_fill, row, "Dockets Booked by Reporter", "Reporter (Booked By)", reporter_table, wt_cols, reporter_names)
    row = _write_summary_table(ws, header_fill, row, "Dockets Cleared by Person", "Cleared By", cleared_table, wt_cols2, cleared_names, cleared_aow, avg_time_cols)
    row = _write_summary_table(ws, header_fill, row, "Pending Dockets by SSA", "SSA_NAME", pending_table, wt_cols3, extra_cols=pending_aow)

    widths = [24] + [16] * (max_value_cols + 1)
    for j, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = width

    return ws, row


def _exclude_reporter_dltl_mask(df):
    """CPAN-DL and MAAN-TL 'is down' dockets booked by EXCLUDE_REPORTER get their
    Date Submitted mis-recorded by GCS as 00:00 hrs of the booking day regardless
    of the actual time booked, which corrupts the Final Report's day/night-duty
    split and pending-aging figures (a docket really booked at 11:13 falls into
    the wrong shift's night-duty bucket). Identifies them so a second Final
    Report can be computed with them excluded, for comparison against manual
    data before this becomes the primary version."""
    is_excluded_reporter = df["Reporter (Booked By)"] == EXCLUDE_REPORTER
    dl_down = _keyword_mask(df, ["CPAN", "DL", "is down"])
    tl_down = _keyword_mask(df, ["MAAN", "TL", "is down"])
    return is_excluded_reporter & (dl_down | tl_down)


def _compute_final_table(df, start_dt, end_dt, cpan_roster=None):
    """Fully reproduces the DAILY_CPAN_DOCKET_CALCULATION reference workbook's
    FINAL-TABLE — including the rows that were manually typed in by hand there
    (day/night duty split, pending-aging buckets) — computed automatically
    from real timestamps instead.

    "Total Booked" and the pending/night-duty rows are scoped to dockets
    actually booked in this shift window. "Total Resolved" / Amount of Work
    Done / Daytime / Night Resolved are scoped to dockets CLEARED in this
    window regardless of when they were booked — so old backlog that finally
    gets resolved during this shift is credited to it.

    cpan_roster: if given (see _load_cpan_staff_roster), a docket cleared by
    someone NOT in the roster is treated as not resolved at all - it falls
    back into "Total Pending Dockets" instead of "Total Resolved", since only
    CPAN staff can validly resolve a docket. None means no filtering (count
    everyone), matching behavior before this feature existed."""
    night_cutoff = start_dt.replace(hour=NIGHT_START_HOUR, minute=0, second=0)
    submitted = pd.to_datetime(df["Date Submitted"], format="%Y-%m-%d %H:%M", errors="coerce")
    cleared = pd.to_datetime(df["Cleared On"], format="%Y-%m-%d %H:%M", errors="coerce")
    is_creation = df["Work Type"] == "Service Creation"
    is_maintenance = df["Work Type"] == "Maintenance"
    booked_mask = df["Booked This Window"]
    resolved_mask = df["Cleared This Window"]  # cleared within window, incl. old backlog
    if cpan_roster is not None:
        resolved_mask = resolved_mask & df["Cleared By"].isin(cpan_roster)
    is_feedback = df["Status"].str.contains("feedback", case=False, na=False)

    def counts(mask):
        return int((mask & is_creation).sum()), int((mask & is_maintenance).sum())

    def work_sum(mask):
        return (
            int(df.loc[mask & is_creation, "Amount of Work"].sum()),
            int(df.loc[mask & is_maintenance, "Amount of Work"].sum()),
        )

    total_c, total_m = counts(booked_mask)
    resolved_c, resolved_m = counts(resolved_mask)
    resolved_work_c, resolved_work_m = work_sum(resolved_mask)
    feedback_c, feedback_m = counts(booked_mask & is_feedback)

    # "Total Resolved" mixes two different populations: dockets booked AND
    # cleared this window, plus older backlog booked earlier that finally got
    # cleared this window. Split them out so "Total - Total Resolved" doesn't
    # look like it should equal "Total Pending" (it doesn't, when backlog is
    # in the mix) - this was a recurring point of confusion in the report.
    resolved_from_today_c, resolved_from_today_m = counts(resolved_mask & booked_mask)
    resolved_backlog_c, resolved_backlog_m = counts(resolved_mask & ~booked_mask)

    daytime_mask = resolved_mask & (cleared < night_cutoff)
    daytime_c, daytime_m = counts(daytime_mask)

    night_resolved_mask = resolved_mask & (cleared >= night_cutoff)
    night_resolved_c, night_resolved_m = counts(night_resolved_mask)

    pending_before_night_mask = booked_mask & (submitted < night_cutoff) & (~resolved_mask | (cleared >= night_cutoff))
    pending_before_c, pending_before_m = counts(pending_before_night_mask)

    booked_after_night_mask = booked_mask & (submitted >= night_cutoff)
    booked_after_c, booked_after_m = counts(booked_after_night_mask)

    total_night_c = pending_before_c + booked_after_c
    total_night_m = pending_before_m + booked_after_m

    now = pd.Timestamp.now()
    age_days = (now - submitted).dt.total_seconds() / 86400
    still_pending = booked_mask & ~resolved_mask

    def aging(lo, hi):
        mask = still_pending & (age_days >= lo)
        if hi is not None:
            mask &= age_days < hi
        return counts(mask)

    p_lt1_c, p_lt1_m = aging(0, 1)
    p_1to3_c, p_1to3_m = aging(1, 3)
    p_gt3_c, p_gt3_m = aging(3, None)

    return {
        "total": (total_c, total_m),
        "resolved": (resolved_c, resolved_m),
        "resolved_work": (resolved_work_c, resolved_work_m),
        "resolved_from_today": (resolved_from_today_c, resolved_from_today_m),
        "resolved_backlog": (resolved_backlog_c, resolved_backlog_m),
        "feedback": (feedback_c, feedback_m),
        "daytime_resolved": (daytime_c, daytime_m),
        "pending_before_night": (pending_before_c, pending_before_m),
        "booked_after_night": (booked_after_c, booked_after_m),
        "total_night": (total_night_c, total_night_m),
        "night_resolved": (night_resolved_c, night_resolved_m),
        "pending_lt1": (p_lt1_c, p_lt1_m),
        "pending_1to3": (p_1to3_c, p_1to3_m),
        "pending_gt3": (p_gt3_c, p_gt3_m),
    }


def _write_final_report_sheet(writer, df, start_dt, end_dt, cpan_roster=None):
    m = _compute_final_table(df, start_dt, end_dt, cpan_roster=cpan_roster)
    ws = writer.book.create_sheet("Final Report")
    ws.sheet_properties.tabColor = HEADER_COLORS["Final Report"]

    title_fill = PatternFill(start_color=FINAL_TITLE_BG, end_color=FINAL_TITLE_BG, fill_type="solid")
    header_fill = PatternFill(start_color=FINAL_HEADER_BG, end_color=FINAL_HEADER_BG, fill_type="solid")
    green_fill = PatternFill(start_color=FINAL_ROW_BG, end_color=FINAL_ROW_BG, fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    pink_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    red_bold = Font(bold=True, color="FF0000")
    magenta_bold = Font(bold=True, color=FINAL_FEEDBACK_TEXT)

    def styled(row, col, value, fill=None, font=None, align=None):
        cell = ws.cell(row=row, column=col, value=value)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        cell.border = border
        return cell

    # Title bar
    ws.merge_cells("A1:G2")
    styled(1, 1, f"Pending Dockets Status: {start_dt:%d-%m-%Y} 08:00 hrs to {end_dt:%d-%m-%Y} 08:00 hrs",
           fill=title_fill, font=title_font, align=center)

    # Table A: "Total Resolved / Work done" — Creation/Maintenance/Total, each
    # split into No. of Dockets / No. of Activity (mirrors reference row 8).
    ws.merge_cells("B4:C4")
    ws.merge_cells("D4:E4")
    ws.merge_cells("F4:G4")
    ws.merge_cells("A4:A5")
    styled(4, 1, "", fill=header_fill)
    for col, label in ((2, "Creation"), (4, "Maintenance"), (6, "Total")):
        styled(4, col, label, fill=header_fill, font=bold, align=center)
    sub_headers = ["No. of\nDockets", "No. of\nActivity"] * 3
    for j, h in enumerate(sub_headers, 2):
        styled(5, j, h, fill=header_fill, font=bold, align=wrap_center)

    styled(6, 1, "Total Resolved / Work done", fill=green_fill, font=bold)
    rc, rm = m["resolved"]
    wc, wm = m["resolved_work"]
    for j, v in enumerate([rc, wc, rm, wm, rc + rm, wc + wm], 2):
        styled(6, j, v, fill=green_fill, align=center)

    # Table B: simple Creation/Maintenance/Total rows (mirrors reference rows 9-21).
    header_row_b = 8
    styled(header_row_b, 1, "", fill=header_fill)
    for j, label in enumerate(["Creation", "Maintenance", "Total"], 2):
        styled(header_row_b, j, label, fill=header_fill, font=bold, align=center)

    def row_vals(label, pair, row, fill=None, font=bold):
        c, mv = pair
        styled(row, 1, label, fill=fill, font=font)
        styled(row, 2, c, fill=fill, align=center)
        styled(row, 3, mv, fill=fill, align=center)
        styled(row, 4, c + mv, fill=fill, align=center)

    sub_font = Font(italic=True, size=10, color="555555")

    row_vals("Today's Booking", m["total"], 9, fill=green_fill)
    row_vals("Total Resolved", m["resolved"], 10, fill=green_fill)
    row_vals("    of which: from Today's Booking", m["resolved_from_today"], 11, font=sub_font)
    row_vals("    of which: Old Backlog Cleared Today", m["resolved_backlog"], 12, font=sub_font)
    row_vals("Returned in Feedback", m["feedback"], 13, font=magenta_bold)
    row_vals("Dockets Resolved in Daytime", m["daytime_resolved"], 14, fill=green_fill)
    row_vals(f"Dockets Pending at {start_dt:%d-%m-%Y} 08:00 PM", m["pending_before_night"], 15)
    row_vals("Dockets Booked after 08:00 PM", m["booked_after_night"], 16)
    row_vals("Total Dockets for Night Duty", m["total_night"], 17, fill=yellow_fill)
    row_vals("Dockets Resolved in Night Duty", m["night_resolved"], 18, fill=green_fill)
    row_vals("Pending  < 1 Day", m["pending_lt1"], 19, font=red_bold)
    row_vals("Pending  1 - 3 days", m["pending_1to3"], 20, font=red_bold)
    row_vals("Pending  > 3 Days", m["pending_gt3"], 21, font=red_bold)

    total_pending_c = m["pending_lt1"][0] + m["pending_1to3"][0] + m["pending_gt3"][0]
    total_pending_m = m["pending_lt1"][1] + m["pending_1to3"][1] + m["pending_gt3"][1]
    row_vals("Total Pending Dockets", (total_pending_c, total_pending_m), 22, fill=pink_fill, font=red_bold)

    # KPI: average clear time per docket (Date Submitted -> Cleared On), among
    # everything cleared this window (incl. old backlog) — staff-performance metric.
    cleared_df = df[df["Cleared This Window"]]
    if cpan_roster is not None:
        cleared_df = cleared_df[cleared_df["Cleared By"].isin(cpan_roster)]
    avg_c = _format_duration_minutes(_avg_clear_minutes(cleared_df, "Service Creation"))
    avg_m = _format_duration_minutes(_avg_clear_minutes(cleared_df, "Maintenance"))
    avg_all = _format_duration_minutes(_avg_clear_minutes(cleared_df))
    kpi_fill = PatternFill(start_color=FINAL_HEADER_BG, end_color=FINAL_HEADER_BG, fill_type="solid")
    styled(24, 1, "Avg Time per Docket (KPI)", fill=kpi_fill, font=bold)
    styled(24, 2, avg_c, fill=kpi_fill, align=center)
    styled(24, 3, avg_m, fill=kpi_fill, align=center)
    styled(24, 4, avg_all, fill=kpi_fill, align=center)

    ws.cell(row=26, column=1, value=(
        "Note: All figures computed automatically from GCS timestamps - no manual entry. "
        f"Day duty ends {NIGHT_START_HOUR}:00; Night duty {NIGHT_START_HOUR}:00-08:00. "
        "Pending-aging is measured from Date Submitted to report-generation time, scoped "
        "to dockets booked within this shift window only. Avg Time per Docket covers all "
        "dockets cleared this window, including old backlog resolved during this shift. "
        "'Total Resolved' = booked-and-cleared-today + old-backlog-cleared-today (see the "
        "two sub-rows below it) - so 'Today's Booking minus Total Resolved' is not the same "
        "number as 'Total Pending Dockets'; use the 'from Today's Booking' sub-row for that comparison."
    )).font = Font(italic=True, size=9, color="0000FF")

    widths = [40, 12, 12, 12, 12, 12, 12]
    for j, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = width


def _add_name_comments(ws, df, columns, id_col, name_col):
    if id_col not in columns or name_col not in df.columns:
        return
    col_idx = columns.index(id_col) + 1
    for row_idx, name in enumerate(df[name_col].tolist(), start=2):
        if name:
            ws.cell(row=row_idx, column=col_idx).comment = Comment(name, "GCS")


def _write_sheet(writer, df, sheet_name):
    display = df[COLUMNS]
    display.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]

    header_color = HEADER_COLORS.get(sheet_name, "444444")
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws.sheet_properties.tabColor = header_color

    for i, col in enumerate(COLUMNS, 1):
        width = max(12, min(50, display[col].astype(str).map(len).max() if len(display) else 12, len(col) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width
        header_cell = ws.cell(row=1, column=i)
        header_cell.fill = header_fill
        header_cell.font = header_font
        if col in TEXT_COLUMNS:
            # openpyxl auto-detects digit-only strings as numbers, which strips
            # leading zeros from IDs (e.g. "0075477" -> 75477); force text cells.
            for row_idx, val in enumerate(display[col], start=2):
                cell = ws.cell(row=row_idx, column=i)
                cell.value = str(val)
                cell.number_format = "@"
    ws.freeze_panes = "A2"

    if "Severity" in COLUMNS:
        sev_idx = COLUMNS.index("Severity") + 1
        for row_idx, val in enumerate(display["Severity"], start=2):
            fill_color = SEVERITY_FILL_COLORS.get(str(val).strip().lower())
            if fill_color:
                ws.cell(row=row_idx, column=sev_idx).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

    _add_name_comments(ws, df, COLUMNS, "Reporter (Booked By)", "Reporter Name")
    _add_name_comments(ws, df, COLUMNS, "Cleared By", "Cleared By Name")


def _write_report_v2(out_path, df, start_dt, end_dt, cpan_roster=None, log=print):
    """Reporter=EXCLUDE_REPORTER is excluded from the Dockets sheet and its
    CPAN population, but their CPAN-DL and MAAN-TL dockets are both reinstated
    into Other (and the MAAN-TL ones additionally into MAAN) — visible in the
    breakdown sheets without polluting the main Dockets view. Summary reflects
    all dockets and adds an extra table for CPAN's own (minor-severity,
    non-EXCLUDE_REPORTER) workload.

    The Final Report sheet's day/night-duty split and pending-aging figures
    additionally exclude EXCLUDE_REPORTER's CPAN-DL/MAAN-TL 'is down' dockets,
    since GCS mis-records their Date Submitted as 00:00 hrs regardless of
    actual booking time - confirmed against manually-verified data. Those
    dockets still appear normally in the Other/MAAN sheets below; only the
    Final Report's own calculation excludes them.

    cpan_roster: if given, "Total Resolved" (Final Report) and "Cleared This
    Window"/"Dockets Cleared by Person" (Summary) only count dockets cleared
    by roster IDs - anyone else's clearance falls back into "pending", since
    only CPAN staff can validly resolve a docket. None = no filtering."""
    dltl_mask = _exclude_reporter_dltl_mask(df)
    df_for_final_report = df[~dltl_mask].reset_index(drop=True)

    df_main = df[df["Reporter (Booked By)"] != EXCLUDE_REPORTER].reset_index(drop=True)
    df_excluded = df[df["Reporter (Booked By)"] == EXCLUDE_REPORTER]

    cpan_pattern = _keyword_mask(df_main, ["CPAN", "DL", "is down"])
    maan_pattern = _keyword_mask(df_main, ["MAAN", "TL", "is down"])
    major_mask = df_main["Severity"].str.lower() == "major"
    other_mask = cpan_pattern | maan_pattern | major_mask
    cpan_df = df_main[~other_mask].reset_index(drop=True)
    other_from_main = df_main[other_mask]

    excluded_cpan = df_excluded[_keyword_mask(df_excluded, ["CPAN", "DL", "is down"])]
    excluded_maan = df_excluded[_keyword_mask(df_excluded, ["MAAN", "TL", "is down"])]

    maan_df = excluded_maan.reset_index(drop=True)
    # Other holds BOTH excluded patterns (CPAN-DL and MAAN-TL), not just CPAN-DL.
    other_df = pd.concat([other_from_main, excluded_cpan, excluded_maan], ignore_index=True)
    sheets = [("CPAN", cpan_df), ("MAAN", maan_df), ("Other", other_df)]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        _write_final_report_sheet(writer, df_for_final_report, start_dt, end_dt, cpan_roster=cpan_roster)
        _write_sheet(writer, df_main, "Dockets")
        for name, sheet_df in sheets:
            _write_sheet(writer, sheet_df, name)

        summary_ws, next_row = _write_summary_sheet(writer, df, cpan_roster=cpan_roster)  # reflects all dockets, incl. EXCLUDE_REPORTER
        header_fill = PatternFill(start_color=HEADER_COLORS["Summary"], end_color=HEADER_COLORS["Summary"], fill_type="solid")
        cpan_staff_table, cpan_staff_cats, cpan_staff_aow = _pivot_summary_with_work(cpan_df, "SSA_NAME", "Work Type")
        _write_summary_table(
            summary_ws, header_fill, next_row,
            "CPAN Docket Clearance by CPAN Staff (SSA Wise)",
            "SSA_NAME", cpan_staff_table, cpan_staff_cats, extra_cols=cpan_staff_aow,
        )

    log(f"Saved {len(df_main)} dockets to {os.path.basename(out_path)} "
        f"({len(df) - len(df_main)} excluded from Dockets, Summary reflects all {len(df)}; "
        f"{int(dltl_mask.sum())} DL/TL-down dockets from reporter {EXCLUDE_REPORTER} additionally "
        f"excluded from the Final Report's day/night-duty and pending calculations)")
    log("  " + ", ".join(f"{name}: {len(sheet_df)}" for name, sheet_df in sheets))


def generate_docket_report(date_str, output_path, log=print):
    """date_str: DD-MM-YYYY, the shift day's start date (covers that day 08:00
    through the next day 08:00). Writes the organization's report to
    output_path, with a "Final Report" sheet first (matches the
    DAILY_CPAN_DOCKET_CALCULATION reference workbook's FINAL-TABLE, computed
    automatically): Reporter=99412764 excluded from the Dockets sheet; their
    CPAN-DL dockets go to Other, MAAN-TL dockets go to MAAN; CPAN sheet is
    clean minor-severity, non-99412764 dockets only. Summary reflects all
    dockets. The Final Report's own day/night-duty split and pending-aging
    additionally exclude reporter 99412764's CPAN-DL/MAAN-TL 'is down'
    dockets, since GCS mis-records their Date Submitted as 00:00 hrs
    regardless of actual booking time (confirmed against manually-verified
    data) - those dockets still appear normally in Other/MAAN.

    If CPAN_STAFF_FILE (cpan_staff_ids.txt) has staff IDs listed, "Total
    Resolved" and "Dockets Cleared by Person" only count dockets cleared by
    those IDs - a docket cleared by anyone else falls back into "pending",
    since only CPAN staff can validly resolve a docket. If the file is empty
    or missing, no filtering is applied (counts everyone).

    Returns output_path."""
    username = os.environ["GCS_USERNAME"]
    password = os.environ["GCS_PASSWORD"]
    login_url = os.environ["GCS_LOGIN_URL"]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    start_dt, end_dt = shift_window(date_str)
    df = _scrape_dockets(date_str, username, password, login_url, log=log)

    cpan_roster = _load_cpan_staff_roster()
    if cpan_roster is not None:
        log(f"CPAN staff roster loaded ({len(cpan_roster)} IDs) - Total Resolved / Cleared by "
            f"Person will only count dockets cleared by these IDs.")
    else:
        log("No CPAN staff roster configured (cpan_staff_ids.txt empty/missing) - "
            "counting all resolvers as before.")

    log(f"Writing report (Reporter={EXCLUDE_REPORTER} excluded from Dockets/CPAN; "
        f"their CPAN-DL dockets reinstated into Other, MAAN-TL reinstated into MAAN)...")
    _write_report_v2(output_path, df, start_dt, end_dt, cpan_roster=cpan_roster, log=log)

    return output_path
