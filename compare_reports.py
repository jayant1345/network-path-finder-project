"""Compare reference report vs our generated report for 02.04.26."""
import openpyxl
import pandas as pd
import sys

REF_PATH = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_02.04.26.xlsx"
GEN_PATH = r"C:\Project_AI\network-path-finder-project\reports_output\DAILY_CPAN_REPORTS_02.04.26.xlsx"

def inspect_sheet(wb, sh_name, label):
    if sh_name not in wb.sheetnames:
        print(f"  [{label}] Sheet '{sh_name}' NOT FOUND")
        return
    sh = wb[sh_name]
    print(f"  [{label}] {sh_name}: max_row={sh.max_row}, max_col={sh.max_column}")
    for r in range(1, min(6, sh.max_row + 1)):
        row_vals = [str(c.value)[:25] if c.value is not None else "-" for c in sh[r]]
        print(f"    Row {r}: {row_vals[:12]}")
    print()

print("=" * 70)
print("REFERENCE REPORT:", REF_PATH)
print("=" * 70)
ref_wb = openpyxl.load_workbook(REF_PATH, data_only=True)
print(f"Sheets: {ref_wb.sheetnames}\n")
for sh_name in ref_wb.sheetnames:
    inspect_sheet(ref_wb, sh_name, "REF")

print("=" * 70)
print("GENERATED REPORT:", GEN_PATH)
print("=" * 70)
gen_wb = openpyxl.load_workbook(GEN_PATH, data_only=True)
print(f"Sheets: {gen_wb.sheetnames}\n")
for sh_name in gen_wb.sheetnames:
    inspect_sheet(gen_wb, sh_name, "GEN")

print("=" * 70)
print("ROW COUNT COMPARISON (data rows only, excluding headers)")
print("=" * 70)
# Check input file row counts
input_files = {
    "CARD OFFLINE": r"C:\Project_AI\cpan_report\CARD OFFLINE 02-04-2026.xlsx",
    "DEVICE OFFLINE": r"C:\Project_AI\cpan_report\DEVICE OFFLINE 02-04-2026.xlsx",
    "FAN FAILURE": r"C:\Project_AI\cpan_report\FAN FAILURE 02-04-2026.xlsx",
    "DL FAIL": r"C:\Project_AI\cpan_report\CPAN DL FAIL REPORT  02-04-2026.xlsx",
}
print("\nINPUT FILES:")
for name, path in input_files.items():
    try:
        df = pd.read_excel(path)
        print(f"  {name}: {len(df)} rows, cols={list(df.columns)[:8]}")
    except Exception as e:
        print(f"  {name}: ERROR - {e}")
