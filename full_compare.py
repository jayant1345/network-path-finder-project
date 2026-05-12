"""
Full comparison: our generated report vs reference (master template output sheets)
for all 5 date folders.
"""
import sys, os, re
sys.path.insert(0, r"C:\Project_AI\network-path-finder-project")

import pandas as pd
import openpyxl
from datetime import datetime
from report_gen import (read_card_off, read_device_off, read_fan_fail,
                        read_dl_fail, read_dash_down)

BASE = r"C:\Project_AI\network-path-finder-project"

# ── Date folder definitions ─────────────────────────────────────────────────
FOLDERS = {
    "05.05.26": {
        "date":       datetime(2026, 5, 5),
        "card_off":   "CARD OFFLINE 05-05-2026.xlsx",
        "device_off": "DEVICE OFFLINE 05-05-2026.xlsx",
        "fan_fail":   "FAN FAILURE 05-05-2026.xlsx",
        "dl_fail":    "CPAN DL FAIL REPORT 05-05-2026.xlsx",
        "dash_down":  "DASH-DOWN_11052026.csv",
        "master":     "DAILY_CPAN_REPORTS_MASTER_05.05.2026.xlsx",
    },
    "06.06.26": {
        "date":       datetime(2026, 5, 6),
        "card_off":   "CARD OFFLINE 06-05-2026.xlsx",
        "device_off": "DEVICE OFFLINE 06-05-2026.xlsx",
        "fan_fail":   "FAN FAILURE 06-05-2026.xlsx",
        "dl_fail":    "CPAN DL FAIL REPORT  06-05-2026.xlsx",
        "dash_down":  "DASH-DOWN_11052026.csv",
        "master":     "DAILY_CPAN_REPORTS_MASTER_06.05.26.xlsx",
    },
    "07.05.26": {
        "date":       datetime(2026, 5, 7),
        "card_off":   "CARD OFFLINE 07.05.2026.xlsx",
        "device_off": "DEVICE OFFLINE 07.05.2026.xlsx",
        "fan_fail":   "FAN FAILURE 07.05.2026.xlsx",
        "dl_fail":    "DL FAIL REPORT 07.05.2026.xlsx",
        "dash_down":  "DASH-DOWN_11052026.csv",
        "master":     "DAILY_CPAN_REPORTS_MASTER_07.05.26.xlsx",
    },
    "08.05.26": {
        "date":       datetime(2026, 5, 8),
        "card_off":   "CARD OFFLINE 08.05.2026.xlsx",
        "device_off": "DEVICE OFFLINE 08.05.2026.xlsx",
        "fan_fail":   "FAN FAILURE 08.05.2026.xlsx",
        "dl_fail":    "DL FAIL REPORT 08.05.2026.xlsx",
        "dash_down":  "DASH-DOWN_11052026.csv",
        "master":     "DAILY_CPAN_REPORTS_MASTER_08.05.26.xlsx",
    },
    "09.05.26": {
        "date":       datetime(2026, 5, 9),
        "card_off":   "CARD OFFLINE 09.05.2026.xlsx",
        "device_off": "DEVICE OFFLINE 09.05.2026.xlsx",
        "fan_fail":   "FAN FAILURE 09.05.2026.xlsx",
        "dl_fail":    "DL FAIL REPORT 09.05.2026.xlsx",
        "dash_down":  "DASH_DOWN.csv",
        "master":     "DAILY_CPAN_REPORTS_MASTER_09.05.26.xlsx",
    },
}

REF_SHEETS = ['CARD-OFF-R', 'DEVICE-OFF-R', 'FAN-FAIL-R', 'DL-FAIL-R']


def ref_data_rows(wb, sheet_name):
    """Return (header_row_index, data_row_count) from master output sheet."""
    if sheet_name not in wb.sheetnames:
        return None, 0
    sh = wb[sheet_name]
    # Find header row: look for 'Sr' in column A within first 10 rows
    hdr_row = None
    for r in range(1, 11):
        val = sh.cell(row=r, column=1).value
        if val is not None and str(val).strip() == 'Sr':
            hdr_row = r
            break
    if hdr_row is None:
        return None, sh.max_row
    # Count non-empty data rows after header
    count = 0
    for r in range(hdr_row + 1, sh.max_row + 1):
        val = sh.cell(row=r, column=1).value
        if val is not None and str(val).strip() not in ('', 'None'):
            count += 1
    return hdr_row, count


def ref_columns(wb, sheet_name, hdr_row):
    """Return column headers from reference sheet."""
    if sheet_name not in wb.sheetnames or hdr_row is None:
        return []
    sh = wb[sheet_name]
    return [str(sh.cell(row=hdr_row, column=c).value or '').strip()
            for c in range(1, sh.max_column + 1)
            if sh.cell(row=hdr_row, column=c).value]


def ref_sample_rows(wb, sheet_name, hdr_row, n=3):
    """Return first n data rows from reference sheet."""
    if sheet_name not in wb.sheetnames or hdr_row is None:
        return []
    sh = wb[sheet_name]
    rows = []
    for r in range(hdr_row + 1, sh.max_row + 1):
        row = [sh.cell(row=r, column=c).value for c in range(1, sh.max_column + 1)]
        if row[0] is not None and str(row[0]).strip() not in ('', 'None'):
            rows.append([str(v)[:30] if v else '-' for v in row[:12]])
            if len(rows) >= n:
                break
    return rows


print("=" * 75)
print("FULL COMPARISON: Generated vs Reference (Master Template Output Sheets)")
print("=" * 75)

summary = []

for folder_name, cfg in FOLDERS.items():
    folder_path = os.path.join(BASE, folder_name)
    report_date = cfg["date"]
    print(f"\n{'='*75}")
    print(f"DATE FOLDER: {folder_name}  ({report_date.strftime('%d-%m-%Y')})")
    print(f"{'='*75}")

    # Load reference master
    master_path = os.path.join(folder_path, cfg["master"])
    try:
        ref_wb = openpyxl.load_workbook(master_path, data_only=True)
        print(f"  Reference master: {cfg['master']}  [Sheets: {ref_wb.sheetnames}]")
    except Exception as e:
        print(f"  ERROR loading master: {e}")
        continue

    # Load DASH-DOWN
    dash_df = None
    dd_path = os.path.join(folder_path, cfg.get("dash_down", ""))
    if dd_path and os.path.exists(dd_path):
        try:
            dash_df = read_dash_down(dd_path)
        except Exception as e:
            print(f"  DASH-DOWN error: {e}")

    readers = {
        'CARD-OFF-R':   lambda p: read_card_off(p, report_date),
        'DEVICE-OFF-R': lambda p: read_device_off(p, report_date, dash_df),
        'FAN-FAIL-R':   lambda p: read_fan_fail(p, report_date),
        'DL-FAIL-R':    lambda p: read_dl_fail(p, report_date),
    }
    input_keys = {
        'CARD-OFF-R':   'card_off',
        'DEVICE-OFF-R': 'device_off',
        'FAN-FAIL-R':   'fan_fail',
        'DL-FAIL-R':    'dl_fail',
    }

    folder_ok = True
    for sheet_name in REF_SHEETS:
        input_file = cfg.get(input_keys[sheet_name], "")
        input_path = os.path.join(folder_path, input_file)

        # Reference count
        hdr_row, ref_count = ref_data_rows(ref_wb, sheet_name)
        ref_cols = ref_columns(ref_wb, sheet_name, hdr_row)

        # Our generated count
        gen_count = 0
        gen_cols  = []
        gen_sample = []
        try:
            df = readers[sheet_name](input_path)
            gen_count  = len(df)
            gen_cols   = list(df.columns)
            gen_sample = df.head(3).values.tolist()
        except Exception as e:
            gen_count = f"ERROR: {e}"

        # Compare
        match = "OK" if isinstance(gen_count, int) and abs(gen_count - ref_count) <= 2 else "MISMATCH"
        if match == "MISMATCH":
            folder_ok = False

        diff = (gen_count - ref_count) if isinstance(gen_count, int) else "?"
        sign = "+" if isinstance(diff, int) and diff > 0 else ""
        print(f"\n  [{match}] {sheet_name}")
        print(f"    REF : {ref_count} rows  | cols: {ref_cols[:8]}")
        print(f"    GEN : {gen_count} rows  | cols: {gen_cols[:8]}  (diff: {sign}{diff})")

        if match == "MISMATCH":
            # Show ref sample
            ref_sample = ref_sample_rows(ref_wb, sheet_name, hdr_row, 2)
            if ref_sample:
                print(f"    REF sample row 1: {ref_sample[0][:8]}")
            if isinstance(gen_count, int) and gen_count > 0 and gen_sample:
                print(f"    GEN sample row 1: {[str(v)[:25] for v in gen_sample[0][:8]]}")

    summary.append((folder_name, "OK" if folder_ok else "MISMATCH"))

print(f"\n{'='*75}")
print("SUMMARY")
print(f"{'='*75}")
for folder_name, status in summary:
    print(f"  {folder_name}: {status}")
all_ok = all(s == "OK" for _, s in summary)
print(f"\nAll folders match: {all_ok}")
