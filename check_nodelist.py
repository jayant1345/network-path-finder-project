"""Check NODE-LIST from master template and understand CARD-OFF-R source logic."""
import openpyxl, pandas as pd

MASTER = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_MASTER_02.04.26.xlsx"
REF    = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_02.04.26.xlsx"
CARD_INPUT = r"C:\Project_AI\cpan_report\CARD OFFLINE 02-04-2026.xlsx"

wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)

print("=== NODE-LIST sheet (first 30 rows) ===")
sh = wb['NODE-LIST']
count = 0
for r in sh.iter_rows(max_row=30, values_only=True):
    if any(v for v in r if v is not None):
        vals = [str(v)[:20] if v is not None else '-' for v in r]
        print(f"  {vals[:10]}")
        count += 1

print(f"\n=== NODE-LIST row count ===")
total_rows = 0
for r in sh.iter_rows(values_only=True):
    if any(v for v in r if v is not None):
        total_rows += 1
print(f"  Non-empty rows in NODE-LIST: {total_rows}")

print()
print("=== Reference CARD-OFF-R: first 10 data rows ===")
ref_wb = openpyxl.load_workbook(REF, data_only=True)
sh_card = ref_wb['CARD-OFF-R']
for r in range(5, 16):
    row = [str(c.value)[:25] if c.value is not None else '-' for c in sh_card[r]]
    print(f"  Row {r}: {row[:10]}")

print()
print("=== Reference CARD-OFF-R: IPs in rows 5-60 ===")
ips_seen = {}
for r in range(5, 60):
    row = [c.value for c in sh_card[r]]
    ip  = str(row[1]).strip() if row[1] else None
    ssa = str(row[5]).strip() if len(row)>5 and row[5] else None
    circ= str(row[4]).strip() if len(row)>4 and row[4] else None
    if ip and ip not in ('None', '-'):
        if ip not in ips_seen:
            ips_seen[ip] = {'ssa': ssa, 'circ': circ, 'count': 0}
        ips_seen[ip]['count'] += 1
print(f"  Unique IPs in rows 5-60: {len(ips_seen)}")
for ip, info in list(ips_seen.items())[:10]:
    print(f"  {ip}: ssa={info['ssa']}, circ={info['circ']}, rows={info['count']}")

print()
print("=== CARD INPUT: GJ IPs and their card count ===")
card_df = pd.read_excel(CARD_INPUT, dtype=str)
card_df.columns = [c.strip() for c in card_df.columns]
gj_df = card_df[card_df['CIRCLE'] == 'GJ']
print(f"GJ rows in input: {len(gj_df)}")
print(f"GJ unique IPs: {gj_df['IP'].nunique()}")
print(f"Cards per IP:")
print(gj_df.groupby('IP')['CARD'].count().value_counts().sort_index().to_string())
print(f"\nSample GJ rows:")
print(gj_df[['IP','NAME','SSA','CARD','Create Time']].head(10).to_string())
