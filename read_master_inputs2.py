"""Read master template input sheets without read_only (slower but reads all data)."""
import openpyxl, pandas as pd

MASTER = r"C:\Project_AI\network-path-finder-project\07.05.26\DAILY_CPAN_REPORTS_MASTER_07.05.26.xlsx"

print("Opening master (full mode, takes longer)...")
wb = openpyxl.load_workbook(MASTER, data_only=True)   # no read_only

def sheet_info(wb, sh_name):
    if sh_name not in wb.sheetnames:
        print(f"\n{sh_name}: NOT FOUND")
        return
    sh = wb[sh_name]
    print(f"\n{'='*60}")
    print(f"{sh_name}: max_row={sh.max_row}, max_col={sh.max_column}")

    # Find first non-empty row
    first_data = None
    for r in range(1, min(20, sh.max_row+1)):
        row = [c.value for c in sh[r]]
        if any(v for v in row if v is not None):
            if first_data is None:
                first_data = r
            print(f"  Row {r}: {[str(v)[:20] if v else '-' for v in row[:10]]}")

    # Count non-empty data rows
    count = 0
    if first_data:
        for r in range(first_data+1, sh.max_row+1):
            row = [c.value for c in sh[r]]
            if any(v for v in row if v is not None):
                count += 1
    print(f"  => ~{count} data rows after header")

for sh_name in ['CARD-OFF', 'DEVICE-OFF', 'FAN-FAIL', 'DL-FAIL', 'DL-EMS']:
    sheet_info(wb, sh_name)
