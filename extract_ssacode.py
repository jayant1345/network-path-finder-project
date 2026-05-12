"""Extract SSA->BA mapping from master template SSACODE sheet."""
import openpyxl, sys

MASTER = r"C:\Project_AI\cpan_report\DAILY_CPAN_REPORTS_MASTER_02.04.26.xlsx"

wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)

for sh_name in ['SSACODE', 'WTR', 'NODE-LIST']:
    if sh_name not in wb.sheetnames:
        print(f"Sheet {sh_name} not found")
        continue
    sh = wb[sh_name]
    print(f"\n=== {sh_name} ===")
    count = 0
    for r in sh.iter_rows(max_row=100, values_only=True):
        if any(v for v in r if v is not None):
            vals = [str(v)[:25] if v is not None else '-' for v in r]
            print(f"  {vals[:10]}")
            count += 1
    print(f"  ... ({count} rows shown, max_row scan to 100)")

# Also try reading SUM-TABLE differently - look for SSA column
print("\n=== SUM-TABLE rows 1-20 ===")
sh_sum = wb['SUM-TABLE']
for r in sh_sum.iter_rows(min_row=1, max_row=20, values_only=True):
    vals = [str(v)[:20] if v is not None else '-' for v in r]
    print(f"  {vals[:10]}")
